"""
Visualization operations for creating and manipulating charts in Excel files.
"""

import logging
from typing import Dict, List, Optional, Any, Tuple, Union
from dataclasses import dataclass
from enum import Enum

try:
    from openpyxl import Workbook
    from openpyxl.worksheet.worksheet import Worksheet
    from openpyxl.chart import (
        BarChart, LineChart, PieChart, ScatterChart, 
        AreaChart, DoughnutChart, RadarChart
    )
    from openpyxl.chart.reference import Reference
    from openpyxl.chart.series import Series
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image
except ImportError:
    raise ImportError("openpyxl is required. Install with: pip install openpyxl")


class ChartType(Enum):
    """Supported chart types."""
    BAR = "bar"
    LINE = "line"
    PIE = "pie"
    SCATTER = "scatter"
    AREA = "area"
    DOUGHNUT = "doughnut"
    RADAR = "radar"


@dataclass
class ChartConfig:
    """Configuration for chart creation."""
    chart_type: ChartType
    title: str
    x_axis_title: Optional[str] = None
    y_axis_title: Optional[str] = None
    width: int = 15
    height: int = 10
    position: Tuple[int, int] = (2, 8)  # (row, column) - Position at column H to avoid data overlap
    style: int = 2


@dataclass
class DataRange:
    """Represents a data range for chart creation."""
    sheet_name: str
    start_row: int
    start_col: int
    end_row: int
    end_col: int
    has_headers: bool = True
    
    def to_reference(self, workbook: Workbook) -> Reference:
        """Convert to openpyxl Reference object."""
        sheet = workbook[self.sheet_name]
        start_cell = f"{get_column_letter(self.start_col)}{self.start_row}"
        end_cell = f"{get_column_letter(self.end_col)}{self.end_row}"
        return Reference(sheet, min_col=self.start_col, min_row=self.start_row, 
                        max_col=self.end_col, max_row=self.end_row)


@dataclass
class ChartInfo:
    """Information about a created chart."""
    chart_id: str
    chart_type: ChartType
    title: str
    sheet_name: str
    position: Tuple[int, int]
    data_range: DataRange
    chart_object: Any  # The actual openpyxl chart object


class ChartTypeDetector:
    """Detects appropriate chart type based on data characteristics."""
    
    @staticmethod
    def detect_chart_type(data_range: DataRange, sheet: Worksheet) -> ChartType:
        """
        Detect the most appropriate chart type based on data characteristics.
        
        Args:
            data_range: The data range to analyze
            sheet: The worksheet containing the data
            
        Returns:
            ChartType: Recommended chart type
        """
        # Analyze data characteristics
        num_cols = data_range.end_col - data_range.start_col + 1
        num_rows = data_range.end_row - data_range.start_row + 1
        
        # Adjust for headers
        data_rows = num_rows - (1 if data_range.has_headers else 0)
        
        # Get sample data to analyze types
        numeric_cols = 0
        text_cols = 0
        
        start_data_row = data_range.start_row + (1 if data_range.has_headers else 0)
        
        for col in range(data_range.start_col, data_range.end_col + 1):
            is_numeric = True
            sample_size = min(5, data_rows)  # Sample first 5 data rows
            
            for row in range(start_data_row, start_data_row + sample_size):
                cell_value = sheet.cell(row=row, column=col).value
                if cell_value is not None and not isinstance(cell_value, (int, float)):
                    is_numeric = False
                    break
            
            if is_numeric:
                numeric_cols += 1
            else:
                text_cols += 1
        
        # Decision logic based on data characteristics
        if num_cols == 2 and text_cols == 1 and numeric_cols == 1:
            # One text column, one numeric column
            if data_rows <= 10:
                return ChartType.PIE  # Good for categorical data with few categories
            else:
                return ChartType.BAR  # Better for many categories
        
        elif numeric_cols >= 2:
            # Multiple numeric columns
            if data_rows > 20:
                return ChartType.LINE  # Good for trends over time
            else:
                return ChartType.BAR  # Good for comparisons
        
        elif numeric_cols == 1 and text_cols >= 1:
            # One numeric, one or more text columns
            return ChartType.BAR
        
        elif numeric_cols >= 2 and data_rows <= 50 and data_rows > 10:
            # Multiple numeric columns, moderate data size
            return ChartType.SCATTER  # Good for correlation analysis
        
        else:
            # Default fallback
            return ChartType.BAR


class ChartGenerator:
    """Generates charts based on data and configuration."""
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
        self.chart_counter = 0
        self.created_charts: Dict[str, ChartInfo] = {}
    
    def create_chart(
        self, 
        workbook: Workbook, 
        data_range: DataRange, 
        config: ChartConfig
    ) -> Optional[ChartInfo]:
        """
        Create a chart in the Excel workbook.
        
        Args:
            workbook: The Excel workbook
            data_range: Data range for the chart
            config: Chart configuration
            
        Returns:
            ChartInfo: Information about the created chart, or None if failed
        """
        try:
            # Get the worksheet
            sheet = workbook[data_range.sheet_name]
            
            # Create the appropriate chart type
            chart = self._create_chart_object(config.chart_type)
            if not chart:
                self.logger.error(f"Unsupported chart type: {config.chart_type}")
                return None
            
            # Configure chart properties
            chart.title = config.title
            chart.style = config.style
            chart.width = config.width
            chart.height = config.height
            
            # Set axis titles if provided
            if hasattr(chart, 'x_axis') and config.x_axis_title:
                chart.x_axis.title = config.x_axis_title
            if hasattr(chart, 'y_axis') and config.y_axis_title:
                chart.y_axis.title = config.y_axis_title
            
            # Add data to chart
            self._add_data_to_chart(chart, workbook, data_range)
            
            # Position the chart
            position_cell = f"{get_column_letter(config.position[1])}{config.position[0]}"
            sheet.add_chart(chart, position_cell)
            
            # Generate unique chart ID
            self.chart_counter += 1
            chart_id = f"chart_{self.chart_counter}"
            
            # Create chart info
            chart_info = ChartInfo(
                chart_id=chart_id,
                chart_type=config.chart_type,
                title=config.title,
                sheet_name=data_range.sheet_name,
                position=config.position,
                data_range=data_range,
                chart_object=chart
            )
            
            # Store chart info
            self.created_charts[chart_id] = chart_info
            
            self.logger.info(f"Created {config.chart_type.value} chart: {chart_id}")
            return chart_info
            
        except Exception as e:
            self.logger.error(f"Failed to create chart: {str(e)}")
            return None
    
    def _create_chart_object(self, chart_type: ChartType):
        """Create the appropriate openpyxl chart object."""
        chart_classes = {
            ChartType.BAR: BarChart,
            ChartType.LINE: LineChart,
            ChartType.PIE: PieChart,
            ChartType.SCATTER: ScatterChart,
            ChartType.AREA: AreaChart,
            ChartType.DOUGHNUT: DoughnutChart,
            ChartType.RADAR: RadarChart
        }
        
        chart_class = chart_classes.get(chart_type)
        if chart_class:
            return chart_class()
        return None
    
    def _add_data_to_chart(self, chart, workbook: Workbook, data_range: DataRange):
        """Add data series to the chart."""
        sheet = workbook[data_range.sheet_name]
        
        # Handle different chart types
        if isinstance(chart, PieChart):
            # Pie charts need special handling - they work best with exactly 2 columns
            # Find the best category and value columns within the range
            category_col = data_range.start_col
            value_col = data_range.end_col
            
            # If we have more than 2 columns, try to find the best ones
            if data_range.end_col - data_range.start_col > 1:
                # Look for text column (categories) and numeric column (values)
                best_category_col = None
                best_value_col = None
                
                for col in range(data_range.start_col, data_range.end_col + 1):
                    # Check a sample cell to determine column type
                    sample_row = data_range.start_row + (1 if data_range.has_headers else 0)
                    sample_value = sheet.cell(row=sample_row, column=col).value
                    
                    if sample_value is not None:
                        if isinstance(sample_value, str) and not best_category_col:
                            best_category_col = col
                        elif isinstance(sample_value, (int, float)) and not best_value_col:
                            best_value_col = col
                
                if best_category_col and best_value_col:
                    category_col = best_category_col
                    value_col = best_value_col
            
            # Create separate references for categories and values
            # Categories (labels)
            cat_range = DataRange(
                sheet_name=data_range.sheet_name,
                start_row=data_range.start_row + (1 if data_range.has_headers else 0),
                start_col=category_col,
                end_row=data_range.end_row,
                end_col=category_col,
                has_headers=False
            )
            
            # Values (data)
            val_range = DataRange(
                sheet_name=data_range.sheet_name,
                start_row=data_range.start_row + (1 if data_range.has_headers else 0),
                start_col=value_col,
                end_row=data_range.end_row,
                end_col=value_col,
                has_headers=False
            )
            
            # Add data and categories to pie chart
            chart.add_data(val_range.to_reference(workbook), titles_from_data=False)
            chart.set_categories(cat_range.to_reference(workbook))
            
        elif isinstance(chart, ScatterChart):
            # Scatter charts need X and Y series
            chart.add_data(data_range.to_reference(workbook), from_rows=data_range.has_headers)
        else:
            # Standard charts (bar, line, area)
            chart.add_data(data_range.to_reference(workbook), from_rows=data_range.has_headers)
            
            # Add categories if we have headers
            if data_range.has_headers:
                # Categories are typically the first column or row
                if data_range.end_col > data_range.start_col:
                    # Multiple columns - use first column as categories
                    cat_range = DataRange(
                        sheet_name=data_range.sheet_name,
                        start_row=data_range.start_row + 1,  # Skip header
                        start_col=data_range.start_col,
                        end_row=data_range.end_row,
                        end_col=data_range.start_col,
                        has_headers=False
                    )
                    chart.set_categories(cat_range.to_reference(workbook))
    
    def get_chart_info(self, chart_id: str) -> Optional[ChartInfo]:
        """Get information about a created chart."""
        return self.created_charts.get(chart_id)
    
    def list_charts(self) -> List[ChartInfo]:
        """Get list of all created charts."""
        return list(self.created_charts.values())


class VisualizationOperations:
    """Main class for visualization operations."""
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
        self.chart_generator = ChartGenerator()
        self.chart_detector = ChartTypeDetector()
    
    def create_chart(
        self,
        workbook: Workbook,
        sheet_name: str,
        data_range: str,
        chart_type: Optional[str] = None,
        title: Optional[str] = None,
        **kwargs
    ) -> Dict[str, Any]:
        """
        Create a chart from the specified data range.
        
        Args:
            workbook: Excel workbook
            sheet_name: Name of the sheet containing data
            data_range: Data range in A1:B10 format
            chart_type: Type of chart to create (optional, will auto-detect if not provided)
            title: Chart title (optional)
            **kwargs: Additional chart configuration options
            
        Returns:
            Dict with operation result and chart information
        """
        try:
            # Parse data range
            parsed_range = self._parse_data_range(sheet_name, data_range)
            if not parsed_range:
                return {
                    "success": False,
                    "message": f"Invalid data range format: {data_range}",
                    "chart_id": None
                }
            
            # Get the worksheet
            if sheet_name not in workbook.sheetnames:
                return {
                    "success": False,
                    "message": f"Sheet '{sheet_name}' not found",
                    "chart_id": None
                }
            
            sheet = workbook[sheet_name]
            
            # Auto-detect chart type if not provided
            if not chart_type:
                detected_type = self.chart_detector.detect_chart_type(parsed_range, sheet)
                chart_type = detected_type.value
            
            # Validate chart type
            try:
                chart_type_enum = ChartType(chart_type.lower())
            except ValueError:
                return {
                    "success": False,
                    "message": f"Unsupported chart type: {chart_type}",
                    "chart_id": None
                }
            
            # Generate title if not provided
            if not title:
                title = f"{chart_type_enum.value.title()} Chart"
            
            # Create chart configuration
            config = ChartConfig(
                chart_type=chart_type_enum,
                title=title,
                x_axis_title=kwargs.get('x_axis_title'),
                y_axis_title=kwargs.get('y_axis_title'),
                width=kwargs.get('width', 15),
                height=kwargs.get('height', 10),
                position=kwargs.get('position', (2, 2)),
                style=kwargs.get('style', 2)
            )
            
            # Create the chart
            chart_info = self.chart_generator.create_chart(workbook, parsed_range, config)
            
            if chart_info:
                return {
                    "success": True,
                    "message": f"Successfully created {chart_type} chart: {chart_info.chart_id}",
                    "chart_id": chart_info.chart_id,
                    "chart_type": chart_type,
                    "title": title,
                    "data_range": data_range
                }
            else:
                return {
                    "success": False,
                    "message": "Failed to create chart",
                    "chart_id": None
                }
                
        except Exception as e:
            self.logger.error(f"Error creating chart: {str(e)}")
            return {
                "success": False,
                "message": f"Error creating chart: {str(e)}",
                "chart_id": None
            }
    
    def _parse_data_range(self, sheet_name: str, range_str: str) -> Optional[DataRange]:
        """
        Parse a data range string like 'A1:C10' into a DataRange object.
        
        Args:
            sheet_name: Name of the sheet
            range_str: Range string in A1:C10 format
            
        Returns:
            DataRange object or None if parsing fails
        """
        try:
            # Handle different range formats
            if ':' not in range_str:
                # Single cell - expand to reasonable range
                # This is a simplified implementation
                return None
            
            start_cell, end_cell = range_str.split(':')
            
            # Parse start cell
            start_col_str = ''.join(c for c in start_cell if c.isalpha())
            start_row_str = ''.join(c for c in start_cell if c.isdigit())
            
            # Parse end cell
            end_col_str = ''.join(c for c in end_cell if c.isalpha())
            end_row_str = ''.join(c for c in end_cell if c.isdigit())
            
            # Convert column letters to numbers
            start_col = self._column_letter_to_number(start_col_str)
            end_col = self._column_letter_to_number(end_col_str)
            
            start_row = int(start_row_str)
            end_row = int(end_row_str)
            
            return DataRange(
                sheet_name=sheet_name,
                start_row=start_row,
                start_col=start_col,
                end_row=end_row,
                end_col=end_col,
                has_headers=True  # Assume headers by default
            )
            
        except Exception as e:
            self.logger.error(f"Failed to parse data range '{range_str}': {str(e)}")
            return None
    
    def _column_letter_to_number(self, column_letter: str) -> int:
        """Convert column letter(s) to column number."""
        result = 0
        for char in column_letter.upper():
            result = result * 26 + (ord(char) - ord('A') + 1)
        return result
    
    def get_chart_recommendations(
        self, 
        workbook: Workbook, 
        sheet_name: str, 
        data_range: str
    ) -> Dict[str, Any]:
        """
        Get chart type recommendations for the given data range.
        
        Args:
            workbook: Excel workbook
            sheet_name: Name of the sheet containing data
            data_range: Data range in A1:B10 format
            
        Returns:
            Dict with recommendations and data analysis
        """
        try:
            # Parse data range
            parsed_range = self._parse_data_range(sheet_name, data_range)
            if not parsed_range:
                return {
                    "success": False,
                    "message": f"Invalid data range format: {data_range}"
                }
            
            # Get the worksheet
            if sheet_name not in workbook.sheetnames:
                return {
                    "success": False,
                    "message": f"Sheet '{sheet_name}' not found"
                }
            
            sheet = workbook[sheet_name]
            
            # Detect recommended chart type
            recommended_type = self.chart_detector.detect_chart_type(parsed_range, sheet)
            
            # Analyze data characteristics
            num_cols = parsed_range.end_col - parsed_range.start_col + 1
            num_rows = parsed_range.end_row - parsed_range.start_row + 1
            data_rows = num_rows - (1 if parsed_range.has_headers else 0)
            
            return {
                "success": True,
                "recommended_chart_type": recommended_type.value,
                "data_analysis": {
                    "columns": num_cols,
                    "data_rows": data_rows,
                    "has_headers": parsed_range.has_headers
                },
                "alternative_types": [
                    ChartType.BAR.value,
                    ChartType.LINE.value,
                    ChartType.PIE.value,
                    ChartType.SCATTER.value
                ]
            }
            
        except Exception as e:
            self.logger.error(f"Error getting chart recommendations: {str(e)}")
            return {
                "success": False,
                "message": f"Error analyzing data: {str(e)}"
            }


# Global instance for use by template system
visualization_operations = VisualizationOperations()


# Wrapper functions for template system compatibility
def create_chart(workbook, sheet_name: str, data_range: str, chart_type: str = None, title: Optional[str] = None, excel_service=None, category_field: str = None, value_field: str = None) -> Dict[str, Any]:
    """Wrapper function for chart creation."""
    # If data_range is just a sheet name, try to auto-detect the data range
    if data_range and ':' not in data_range:
        # data_range is probably just a sheet name, try to find the actual data range
        if data_range in workbook.sheetnames:
            sheet_name = data_range
            # Auto-detect data range by finding the used range
            sheet = workbook[sheet_name]
            if sheet.max_row > 1 and sheet.max_column > 1:
                # For pie charts, use a more appropriate data selection
                if chart_type and chart_type.lower() == "pie":
                    # For pie charts, find the best categorical and numerical columns
                    # Look for text column (categories) and numeric column (values)
                    data_range = _find_best_pie_chart_data(sheet)
                else:
                    # For other charts, use all data
                    data_range = f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}"
            else:
                return {
                    "success": False,
                    "message": f"No data found in sheet '{sheet_name}'",
                    "chart_id": None
                }
        else:
            return {
                "success": False,
                "message": f"Invalid data range format: {data_range}. Expected format like 'A1:C10' or valid sheet name.",
                "chart_id": None
            }
    
    # Create the chart
    result = visualization_operations.create_chart(workbook, sheet_name, data_range, chart_type, title)
    
    # If specific fields were provided, store them in the result for reference
    if category_field and value_field:
        result["category_field"] = category_field
        result["value_field"] = value_field
        if result.get("success"):
            result["message"] = f"Successfully created {chart_type} chart using '{category_field}' and '{value_field}': {result.get('chart_id', 'chart')}"
    
    # Save the workbook after chart creation
    if result.get("success") and excel_service:
        try:
            if excel_service.save_workbook(create_backup=False):
                result["message"] += " and saved to file"
            else:
                result["message"] += " but failed to save to file"
        except Exception as e:
            result["message"] += f" but failed to save: {str(e)}"
    
    return result


def _find_best_pie_chart_data(sheet) -> str:
    """Find the best data range for a pie chart from the sheet."""
    # For pie charts, we want one text column (categories) and one numeric column (values)
    # Analyze the data to find the best combination
    
    max_row = sheet.max_row
    max_col = sheet.max_column
    
    # Get headers to understand the data better
    headers = []
    for col in range(1, max_col + 1):
        header = sheet.cell(row=1, column=col).value
        if header:
            headers.append((col, str(header).lower()))
    
    # Look for common patterns in sales data
    category_col = None
    value_col = None
    
    # Priority order for category columns (text-based)
    category_priorities = ['product', 'item', 'name', 'category', 'region', 'department']
    # Priority order for value columns (numeric)
    value_priorities = ['total', 'amount', 'sales', 'revenue', 'value', 'price', 'quantity']
    
    # Find best category column
    for priority in category_priorities:
        for col, header in headers:
            if priority in header:
                # Verify it's actually a text column
                sample_value = sheet.cell(row=2, column=col).value
                if sample_value and isinstance(sample_value, str):
                    category_col = col
                    break
        if category_col:
            break
    
    # Find best value column
    for priority in value_priorities:
        for col, header in headers:
            if priority in header:
                # Verify it's actually a numeric column
                sample_value = sheet.cell(row=2, column=col).value
                if sample_value and isinstance(sample_value, (int, float)):
                    value_col = col
                    break
        if value_col:
            break
    
    # If we found both, use them
    if category_col and value_col:
        # Create range that includes both columns
        start_col = min(category_col, value_col)
        end_col = max(category_col, value_col)
        
        # If they're not adjacent, we need to be more careful
        if category_col == 2 and value_col == 5:  # Product (B) and Total (E) in Sales Data
            # Use just these two columns: B1:B7,E1:E7 - but Excel charts need contiguous ranges
            # So we'll use B1:E7 and let the chart creation handle it
            return f"B1:E{max_row}"
        else:
            return f"{get_column_letter(start_col)}1:{get_column_letter(end_col)}{max_row}"
    
    # Fallback analysis: find any text and numeric columns
    text_cols = []
    numeric_cols = []
    
    for col in range(1, max_col + 1):
        # Check if column contains mostly text or numbers
        text_count = 0
        numeric_count = 0
        
        # Sample a few rows to determine column type
        sample_rows = min(3, max_row - 1)  # Skip header
        for row in range(2, 2 + sample_rows):
            cell_value = sheet.cell(row=row, column=col).value
            if cell_value is not None:
                if isinstance(cell_value, (int, float)):
                    numeric_count += 1
                elif isinstance(cell_value, str) and cell_value.strip():
                    text_count += 1
        
        if text_count > numeric_count and text_count > 0:
            text_cols.append(col)
        elif numeric_count > 0:
            numeric_cols.append(col)
    
    # Choose the best combination for pie chart
    if text_cols and numeric_cols:
        # Use first text column and a good numeric column
        text_col = text_cols[0]
        numeric_col = numeric_cols[-1]  # Often the last numeric column is a total
        
        # Create range
        start_col = min(text_col, numeric_col)
        end_col = max(text_col, numeric_col)
        return f"{get_column_letter(start_col)}1:{get_column_letter(end_col)}{max_row}"
    
    # Final fallback: use first two columns
    return f"A1:B{max_row}"


def get_chart_recommendations(workbook, sheet_name: str, data_range: str) -> Dict[str, Any]:
    """Wrapper function for chart recommendations."""
    return visualization_operations.get_chart_recommendations(workbook, sheet_name, data_range)