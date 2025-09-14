"""
Data analysis operations for Excel-LLM system.

This module provides statistical analysis and data insights operations.
"""

import logging
from typing import Dict, Any, List, Optional
from statistics import mean, median, mode, stdev
from collections import Counter

def calculate_statistics(workbook, sheet_name: str, column: str, stat_type: str = "all") -> Dict[str, Any]:
    """
    Calculate statistical measures for a column of data.
    
    Args:
        workbook: Excel workbook object
        sheet_name: Name of the target sheet
        column: Column letter (A, B, C) or name to analyze
        stat_type: Type of statistics to calculate (all, mean, median, mode, std)
        
    Returns:
        Dict with success status and statistical results
    """
    try:
        if sheet_name not in workbook.sheetnames:
            return {
                "success": False,
                "message": f"Sheet '{sheet_name}' not found",
                "data": None
            }
        
        sheet = workbook[sheet_name]
        
        # Convert column letter to number if needed
        if len(column) == 1 and column.isalpha():
            from openpyxl.utils import column_index_from_string
            col_num = column_index_from_string(column.upper())
        else:
            # Try to find column by header name
            col_num = None
            for col in range(1, sheet.max_column + 1):
                header = sheet.cell(row=1, column=col).value
                if header and str(header).lower() == column.lower():
                    col_num = col
                    break
            
            if col_num is None:
                return {
                    "success": False,
                    "message": f"Column '{column}' not found",
                    "data": None
                }
        
        # Extract numerical data from column
        values = []
        for row in range(2, sheet.max_row + 1):  # Skip header row
            cell_value = sheet.cell(row=row, column=col_num).value
            if cell_value is not None:
                try:
                    # Try to convert to float
                    if isinstance(cell_value, (int, float)):
                        values.append(float(cell_value))
                    else:
                        # Try to parse string as number
                        values.append(float(str(cell_value)))
                except (ValueError, TypeError):
                    # Skip non-numeric values
                    continue
        
        if not values:
            return {
                "success": False,
                "message": f"No numeric data found in column '{column}'",
                "data": None
            }
        
        # Calculate statistics
        stats = {}
        
        if stat_type in ["all", "mean"]:
            stats["mean"] = round(mean(values), 2)
        
        if stat_type in ["all", "median"]:
            stats["median"] = round(median(values), 2)
        
        if stat_type in ["all", "mode"]:
            try:
                stats["mode"] = mode(values)
            except:
                # No unique mode
                counter = Counter(values)
                most_common = counter.most_common(1)
                stats["mode"] = most_common[0][0] if most_common else None
        
        if stat_type in ["all", "std", "stdev"]:
            if len(values) > 1:
                stats["standard_deviation"] = round(stdev(values), 2)
            else:
                stats["standard_deviation"] = 0
        
        if stat_type in ["all", "min", "max", "range"]:
            stats["min"] = min(values)
            stats["max"] = max(values)
            stats["range"] = round(max(values) - min(values), 2)
        
        if stat_type == "all":
            stats["count"] = len(values)
            stats["sum"] = round(sum(values), 2)
        
        return {
            "success": True,
            "message": f"Statistics calculated for column '{column}' in {sheet_name}",
            "data": {
                "column": column,
                "sample_size": len(values),
                "statistics": stats
            }
        }
        
    except Exception as e:
        logging.error(f"Error calculating statistics: {str(e)}")
        return {
            "success": False,
            "message": f"Error calculating statistics: {str(e)}",
            "data": None
        }


def find_outliers(workbook, sheet_name: str, column: str, method: str = "iqr") -> Dict[str, Any]:
    """
    Find outliers in a column of data.
    
    Args:
        workbook: Excel workbook object
        sheet_name: Name of the target sheet
        column: Column letter (A, B, C) or name to analyze
        method: Method to detect outliers (iqr, zscore)
        
    Returns:
        Dict with success status and outlier information
    """
    try:
        if sheet_name not in workbook.sheetnames:
            return {
                "success": False,
                "message": f"Sheet '{sheet_name}' not found",
                "data": None
            }
        
        sheet = workbook[sheet_name]
        
        # Convert column letter to number if needed
        if len(column) == 1 and column.isalpha():
            from openpyxl.utils import column_index_from_string
            col_num = column_index_from_string(column.upper())
        else:
            # Try to find column by header name
            col_num = None
            for col in range(1, sheet.max_column + 1):
                header = sheet.cell(row=1, column=col).value
                if header and str(header).lower() == column.lower():
                    col_num = col
                    break
            
            if col_num is None:
                return {
                    "success": False,
                    "message": f"Column '{column}' not found",
                    "data": None
                }
        
        # Extract numerical data with row numbers
        values_with_rows = []
        for row in range(2, sheet.max_row + 1):  # Skip header row
            cell_value = sheet.cell(row=row, column=col_num).value
            if cell_value is not None:
                try:
                    if isinstance(cell_value, (int, float)):
                        values_with_rows.append((float(cell_value), row))
                    else:
                        values_with_rows.append((float(str(cell_value)), row))
                except (ValueError, TypeError):
                    continue
        
        if len(values_with_rows) < 4:
            return {
                "success": False,
                "message": f"Not enough numeric data for outlier detection (need at least 4 values)",
                "data": None
            }
        
        values = [v[0] for v in values_with_rows]
        outliers = []
        
        if method.lower() == "iqr":
            # Interquartile Range method
            sorted_values = sorted(values)
            n = len(sorted_values)
            
            q1_idx = n // 4
            q3_idx = 3 * n // 4
            
            q1 = sorted_values[q1_idx]
            q3 = sorted_values[q3_idx]
            iqr = q3 - q1
            
            lower_bound = q1 - 1.5 * iqr
            upper_bound = q3 + 1.5 * iqr
            
            for value, row in values_with_rows:
                if value < lower_bound or value > upper_bound:
                    outliers.append({
                        "value": value,
                        "row": row,
                        "type": "low" if value < lower_bound else "high"
                    })
        
        elif method.lower() == "zscore":
            # Z-score method
            mean_val = mean(values)
            std_val = stdev(values) if len(values) > 1 else 0
            
            if std_val == 0:
                return {
                    "success": False,
                    "message": "Cannot calculate z-score: standard deviation is zero",
                    "data": None
                }
            
            for value, row in values_with_rows:
                z_score = abs((value - mean_val) / std_val)
                if z_score > 2.5:  # Threshold for outlier
                    outliers.append({
                        "value": value,
                        "row": row,
                        "z_score": round(z_score, 2),
                        "type": "high" if value > mean_val else "low"
                    })
        
        return {
            "success": True,
            "message": f"Found {len(outliers)} outliers in column '{column}' using {method.upper()} method",
            "data": {
                "column": column,
                "method": method.upper(),
                "total_values": len(values),
                "outlier_count": len(outliers),
                "outliers": outliers
            }
        }
        
    except Exception as e:
        logging.error(f"Error finding outliers: {str(e)}")
        return {
            "success": False,
            "message": f"Error finding outliers: {str(e)}",
            "data": None
        }


def data_summary(workbook, sheet_name: str) -> Dict[str, Any]:
    """
    Generate a comprehensive data summary for a sheet.
    
    Args:
        workbook: Excel workbook object
        sheet_name: Name of the target sheet
        
    Returns:
        Dict with success status and data summary
    """
    try:
        if sheet_name not in workbook.sheetnames:
            return {
                "success": False,
                "message": f"Sheet '{sheet_name}' not found",
                "data": None
            }
        
        sheet = workbook[sheet_name]
        
        # Basic sheet info
        summary = {
            "sheet_name": sheet_name,
            "total_rows": sheet.max_row,
            "total_columns": sheet.max_column,
            "data_rows": sheet.max_row - 1 if sheet.max_row > 1 else 0,
            "columns": []
        }
        
        # Analyze each column
        for col in range(1, sheet.max_column + 1):
            header = sheet.cell(row=1, column=col).value
            column_name = str(header) if header else f"Column_{col}"
            
            # Count data types
            numeric_count = 0
            text_count = 0
            empty_count = 0
            date_count = 0
            
            values = []
            
            for row in range(2, sheet.max_row + 1):
                cell_value = sheet.cell(row=row, column=col).value
                
                if cell_value is None or cell_value == "":
                    empty_count += 1
                elif isinstance(cell_value, (int, float)):
                    numeric_count += 1
                    values.append(cell_value)
                elif hasattr(cell_value, 'date'):  # datetime object
                    date_count += 1
                else:
                    text_count += 1
                    # Try to parse as number
                    try:
                        float_val = float(str(cell_value))
                        numeric_count += 1
                        text_count -= 1
                        values.append(float_val)
                    except:
                        pass
            
            column_info = {
                "name": column_name,
                "position": col,
                "data_type": "mixed",
                "numeric_count": numeric_count,
                "text_count": text_count,
                "date_count": date_count,
                "empty_count": empty_count,
                "fill_rate": round((sheet.max_row - 1 - empty_count) / max(1, sheet.max_row - 1) * 100, 1)
            }
            
            # Determine primary data type
            total_non_empty = numeric_count + text_count + date_count
            if total_non_empty > 0:
                if numeric_count / total_non_empty > 0.8:
                    column_info["data_type"] = "numeric"
                    if values:
                        column_info["min_value"] = min(values)
                        column_info["max_value"] = max(values)
                        column_info["avg_value"] = round(mean(values), 2)
                elif date_count / total_non_empty > 0.8:
                    column_info["data_type"] = "date"
                elif text_count / total_non_empty > 0.8:
                    column_info["data_type"] = "text"
            
            summary["columns"].append(column_info)
        
        return {
            "success": True,
            "message": f"Data summary generated for {sheet_name}",
            "data": summary
        }
        
    except Exception as e:
        logging.error(f"Error generating data summary: {str(e)}")
        return {
            "success": False,
            "message": f"Error generating data summary: {str(e)}",
            "data": None
        }