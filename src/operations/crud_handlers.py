"""
CRUD operation handlers for Excel data manipulation.

This module provides handlers for Create, Read, Update, and Delete operations
on Excel data with validation, safety checks, and confirmation messaging.
"""

import logging
from typing import Dict, List, Optional, Any, Union, Tuple
from dataclasses import dataclass
from datetime import datetime
import re

from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter, column_index_from_string

from excel.excel_service import ExcelService, ExcelStructure
from safety.safety_manager import SafetyManager

# Add src directory to Python path for imports
import sys
from pathlib import Path
current_dir = Path(__file__).parent
src_dir = current_dir.parent if current_dir.name != 'src' else current_dir
if str(src_dir) not in sys.path:
    sys.path.insert(0, str(src_dir))



@dataclass
class OperationResult:
    """Result of a CRUD operation."""
    success: bool
    message: str
    affected_rows: int = 0
    affected_columns: int = 0
    data: Optional[Any] = None
    warnings: List[str] = None
    
    def __post_init__(self):
        if self.warnings is None:
            self.warnings = []


@dataclass
class InsertionData:
    """Data structure for insertion operations."""
    values: List[Any]
    target_sheet: str
    target_row: Optional[int] = None
    target_column: Optional[str] = None
    column_names: Optional[List[str]] = None
    insert_type: str = "row"  # "row" or "column"


class DataInsertionHandler:
    """Handler for data insertion operations (Create operations)."""
    
    def __init__(self, excel_service: ExcelService, safety_manager: SafetyManager):
        """
        Initialize the data insertion handler.
        
        Args:
            excel_service: Excel service instance
            safety_manager: Safety manager for validation
        """
        self.excel_service = excel_service
        self.safety_manager = safety_manager
        self.logger = logging.getLogger(__name__)
    
    def insert_row(self, data: InsertionData) -> OperationResult:
        """
        Insert a new row of data into the Excel sheet.
        
        Args:
            data: InsertionData containing values and target information
            
        Returns:
            OperationResult: Result of the insertion operation
        """
        try:
            # Validate inputs
            validation_result = self._validate_row_insertion(data)
            if not validation_result.success:
                return validation_result
            
            # Get worksheet
            sheet = self.excel_service.get_sheet(data.target_sheet)
            if not sheet:
                return OperationResult(
                    success=False,
                    message=f"Sheet '{data.target_sheet}' not found"
                )
            
            # Create backup before operation
            backup_path = self.excel_service.create_backup()
            if not backup_path:
                self.logger.warning("Failed to create backup before row insertion")
            
            # Determine insertion row
            if data.target_row is None:
                # Insert at the end of existing data
                target_row = sheet.max_row + 1
            else:
                target_row = data.target_row
                # Shift existing rows down if inserting in the middle
                if target_row <= sheet.max_row:
                    sheet.insert_rows(target_row)
            
            # Insert data
            inserted_count = 0
            for col_idx, value in enumerate(data.values, 1):
                if col_idx > sheet.max_column and sheet.max_column > 0:
                    # Adding new columns beyond existing structure
                    break
                
                cell = sheet.cell(row=target_row, column=col_idx)
                
                # Validate and convert data type
                converted_value = self._validate_and_convert_value(
                    value, data.target_sheet, col_idx
                )
                
                cell.value = converted_value
                inserted_count += 1
            
            # Save workbook
            if not self.excel_service.save_workbook(create_backup=False):
                return OperationResult(
                    success=False,
                    message="Failed to save workbook after row insertion"
                )
            
            # Update structure analysis
            self.excel_service.structure = self.excel_service._analyze_structure()
            
            return OperationResult(
                success=True,
                message=f"Successfully inserted row at position {target_row} with {inserted_count} values",
                affected_rows=1,
                data={"row": target_row, "values_inserted": inserted_count}
            )
            
        except Exception as e:
            self.logger.error(f"Error inserting row: {str(e)}")
            return OperationResult(
                success=False,
                message=f"Failed to insert row: {str(e)}"
            )
    
    def insert_column(self, data: InsertionData) -> OperationResult:
        """
        Insert a new column of data into the Excel sheet.
        
        Args:
            data: InsertionData containing values and target information
            
        Returns:
            OperationResult: Result of the insertion operation
        """
        try:
            # Validate inputs
            validation_result = self._validate_column_insertion(data)
            if not validation_result.success:
                return validation_result
            
            # Get worksheet
            sheet = self.excel_service.get_sheet(data.target_sheet)
            if not sheet:
                return OperationResult(
                    success=False,
                    message=f"Sheet '{data.target_sheet}' not found"
                )
            
            # Create backup before operation
            backup_path = self.excel_service.create_backup()
            if not backup_path:
                self.logger.warning("Failed to create backup before column insertion")
            
            # Determine insertion column
            if data.target_column is None:
                # Insert at the end of existing columns
                target_col = sheet.max_column + 1
            else:
                if data.target_column.isdigit():
                    target_col = int(data.target_column)
                else:
                    target_col = column_index_from_string(data.target_column)
                
                # Shift existing columns right if inserting in the middle
                if target_col <= sheet.max_column:
                    sheet.insert_cols(target_col)
            
            # Insert column header if provided
            header_row = 1
            if data.column_names and len(data.column_names) > 0:
                sheet.cell(row=header_row, column=target_col).value = data.column_names[0]
                start_row = 2
            else:
                start_row = 1
            
            # Insert data values
            inserted_count = 0
            for row_idx, value in enumerate(data.values, start_row):
                if row_idx > sheet.max_row + len(data.values):
                    break
                
                cell = sheet.cell(row=row_idx, column=target_col)
                
                # Validate and convert data type
                converted_value = self._validate_and_convert_value(
                    value, data.target_sheet, target_col
                )
                
                cell.value = converted_value
                inserted_count += 1
            
            # Save workbook
            if not self.excel_service.save_workbook(create_backup=False):
                return OperationResult(
                    success=False,
                    message="Failed to save workbook after column insertion"
                )
            
            # Update structure analysis
            self.excel_service.structure = self.excel_service._analyze_structure()
            
            column_letter = get_column_letter(target_col)
            return OperationResult(
                success=True,
                message=f"Successfully inserted column at position {column_letter} with {inserted_count} values",
                affected_columns=1,
                data={"column": column_letter, "values_inserted": inserted_count}
            )
            
        except Exception as e:
            self.logger.error(f"Error inserting column: {str(e)}")
            return OperationResult(
                success=False,
                message=f"Failed to insert column: {str(e)}"
            )
    
    def add_multiple_rows(self, rows_data: List[InsertionData]) -> OperationResult:
        """
        Add multiple rows of data in a single operation.
        
        Args:
            rows_data: List of InsertionData for each row
            
        Returns:
            OperationResult: Result of the batch insertion operation
        """
        if not rows_data:
            return OperationResult(
                success=False,
                message="No data provided for insertion"
            )
        
        # Safety check: limit batch size
        if len(rows_data) > 50:
            return OperationResult(
                success=False,
                message="Batch insertion limited to 50 rows for safety. Please split into smaller batches."
            )
        
        successful_insertions = 0
        failed_insertions = 0
        error_messages = []
        
        for i, row_data in enumerate(rows_data):
            result = self.insert_row(row_data)
            if result.success:
                successful_insertions += 1
            else:
                failed_insertions += 1
                error_messages.append(f"Row {i+1}: {result.message}")
        
        if failed_insertions == 0:
            return OperationResult(
                success=True,
                message=f"Successfully inserted {successful_insertions} rows",
                affected_rows=successful_insertions
            )
        elif successful_insertions > 0:
            return OperationResult(
                success=True,
                message=f"Partially successful: {successful_insertions} rows inserted, {failed_insertions} failed",
                affected_rows=successful_insertions,
                warnings=error_messages
            )
        else:
            return OperationResult(
                success=False,
                message=f"All {failed_insertions} row insertions failed",
                warnings=error_messages
            )
    
    def _validate_row_insertion(self, data: InsertionData) -> OperationResult:
        """
        Validate row insertion data and parameters.
        
        Args:
            data: InsertionData to validate
            
        Returns:
            OperationResult: Validation result
        """
        if not data.values:
            return OperationResult(
                success=False,
                message="No values provided for row insertion"
            )
        
        if not data.target_sheet:
            return OperationResult(
                success=False,
                message="Target sheet not specified"
            )
        
        # Check if sheet exists
        if data.target_sheet not in self.excel_service.get_sheet_names():
            return OperationResult(
                success=False,
                message=f"Sheet '{data.target_sheet}' does not exist"
            )
        
        # Validate row position
        if data.target_row is not None and data.target_row < 1:
            return OperationResult(
                success=False,
                message="Row position must be greater than 0"
            )
        
        # Safety check: limit number of values per row
        if len(data.values) > 100:
            return OperationResult(
                success=False,
                message="Row insertion limited to 100 values for safety"
            )
        
        return OperationResult(success=True, message="Validation passed")
    
    def _validate_column_insertion(self, data: InsertionData) -> OperationResult:
        """
        Validate column insertion data and parameters.
        
        Args:
            data: InsertionData to validate
            
        Returns:
            OperationResult: Validation result
        """
        if not data.values:
            return OperationResult(
                success=False,
                message="No values provided for column insertion"
            )
        
        if not data.target_sheet:
            return OperationResult(
                success=False,
                message="Target sheet not specified"
            )
        
        # Check if sheet exists
        if data.target_sheet not in self.excel_service.get_sheet_names():
            return OperationResult(
                success=False,
                message=f"Sheet '{data.target_sheet}' does not exist"
            )
        
        # Validate column position
        if data.target_column is not None:
            if data.target_column.isdigit():
                if int(data.target_column) < 1:
                    return OperationResult(
                        success=False,
                        message="Column position must be greater than 0"
                    )
            else:
                try:
                    column_index_from_string(data.target_column)
                except ValueError:
                    return OperationResult(
                        success=False,
                        message=f"Invalid column reference: {data.target_column}"
                    )
        
        # Safety check: limit number of values per column
        if len(data.values) > 1000:
            return OperationResult(
                success=False,
                message="Column insertion limited to 1000 values for safety"
            )
        
        return OperationResult(success=True, message="Validation passed")
    
    def _validate_and_convert_value(self, value: Any, sheet_name: str, column_index: int) -> Any:
        """
        Validate and convert a value to match the expected column data type.
        
        Args:
            value: Value to validate and convert
            sheet_name: Name of the target sheet
            column_index: Column index (1-based)
            
        Returns:
            Any: Converted value
        """
        if value is None or value == "":
            return None
        
        # Get expected data type from structure analysis
        structure = self.excel_service.get_structure()
        if not structure:
            return value
        
        sheet_info = structure.get_sheet_info(sheet_name)
        headers = sheet_info.get('headers', [])
        data_types = sheet_info.get('data_types', {})
        
        # Get column header and expected type
        if column_index <= len(headers):
            header = headers[column_index - 1]
            expected_type = data_types.get(header, 'text')
        else:
            expected_type = 'text'
        
        # Convert based on expected type
        try:
            if expected_type == 'number':
                if isinstance(value, str):
                    # Try to convert string to number
                    if '.' in value:
                        return float(value)
                    else:
                        return int(value)
                elif isinstance(value, (int, float)):
                    return value
            elif expected_type == 'boolean':
                if isinstance(value, str):
                    return value.lower() in ('true', '1', 'yes', 'on')
                elif isinstance(value, bool):
                    return value
            elif expected_type == 'date':
                if isinstance(value, str):
                    # Try to parse common date formats
                    for fmt in ['%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y']:
                        try:
                            return datetime.strptime(value, fmt)
                        except ValueError:
                            continue
                elif isinstance(value, datetime):
                    return value
        except (ValueError, TypeError):
            # If conversion fails, return as string
            pass
        
        return str(value)


@dataclass
class QueryData:
    """Data structure for query operations."""
    target_sheet: str
    columns: Optional[List[str]] = None  # Columns to select, None for all
    conditions: Optional[Dict[str, Any]] = None  # Filter conditions
    sort_by: Optional[str] = None  # Column to sort by
    sort_order: str = "asc"  # "asc" or "desc"
    limit: Optional[int] = None  # Limit number of results
    aggregations: Optional[Dict[str, str]] = None  # Column -> function mapping


@dataclass
class QueryResult:
    """Result of a query operation."""
    success: bool
    message: str
    data: Optional[List[Dict[str, Any]]] = None
    headers: Optional[List[str]] = None
    row_count: int = 0
    aggregation_results: Optional[Dict[str, Any]] = None


class DataQueryHandler:
    """Handler for data query and read operations (Read operations)."""
    
    def __init__(self, excel_service: ExcelService, safety_manager: SafetyManager):
        """
        Initialize the data query handler.
        
        Args:
            excel_service: Excel service instance
            safety_manager: Safety manager for validation
        """
        self.excel_service = excel_service
        self.safety_manager = safety_manager
        self.logger = logging.getLogger(__name__)
    
    def query_data(self, query: QueryData) -> QueryResult:
        """
        Execute a data query on the Excel sheet.
        
        Args:
            query: QueryData containing query parameters
            
        Returns:
            QueryResult: Result of the query operation
        """
        try:
            # Validate query
            validation_result = self._validate_query(query)
            if not validation_result.success:
                return QueryResult(
                    success=False,
                    message=validation_result.message
                )
            
            # Get worksheet
            sheet = self.excel_service.get_sheet(query.target_sheet)
            if not sheet:
                return QueryResult(
                    success=False,
                    message=f"Sheet '{query.target_sheet}' not found"
                )
            
            # Get sheet structure
            structure = self.excel_service.get_structure()
            if not structure:
                return QueryResult(
                    success=False,
                    message="Unable to analyze sheet structure"
                )
            
            sheet_info = structure.get_sheet_info(query.target_sheet)
            headers = sheet_info.get('headers', [])
            
            if not headers:
                return QueryResult(
                    success=False,
                    message="No headers found in sheet"
                )
            
            # Extract data from sheet
            raw_data = self._extract_sheet_data(sheet, headers)
            
            # Apply filters
            filtered_data = self._apply_filters(raw_data, query.conditions)
            
            # Apply column selection
            selected_data = self._select_columns(filtered_data, query.columns, headers)
            
            # Apply sorting
            sorted_data = self._apply_sorting(selected_data, query.sort_by, query.sort_order)
            
            # Apply limit
            limited_data = self._apply_limit(sorted_data, query.limit)
            
            # Calculate aggregations if requested
            aggregation_results = None
            if query.aggregations:
                aggregation_results = self._calculate_aggregations(filtered_data, query.aggregations)
            
            # Determine final headers
            final_headers = query.columns if query.columns else headers
            
            return QueryResult(
                success=True,
                message=f"Query executed successfully. Retrieved {len(limited_data)} rows.",
                data=limited_data,
                headers=final_headers,
                row_count=len(limited_data),
                aggregation_results=aggregation_results
            )
            
        except Exception as e:
            self.logger.error(f"Error executing query: {str(e)}")
            return QueryResult(
                success=False,
                message=f"Failed to execute query: {str(e)}"
            )
    
    def get_sheet_summary(self, sheet_name: str) -> QueryResult:
        """
        Get a summary of the sheet data including basic statistics.
        
        Args:
            sheet_name: Name of the sheet to summarize
            
        Returns:
            QueryResult: Summary information
        """
        try:
            structure = self.excel_service.get_structure()
            if not structure:
                return QueryResult(
                    success=False,
                    message="Unable to analyze sheet structure"
                )
            
            sheet_info = structure.get_sheet_info(sheet_name)
            
            summary = {
                "sheet_name": sheet_name,
                "total_rows": sheet_info.get('row_count', 0),
                "total_columns": sheet_info.get('column_count', 0),
                "headers": sheet_info.get('headers', []),
                "data_types": sheet_info.get('data_types', {})
            }
            
            # Get sample data (first 5 rows)
            sheet = self.excel_service.get_sheet(sheet_name)
            if sheet:
                headers = sheet_info.get('headers', [])
                sample_data = self._extract_sheet_data(sheet, headers, max_rows=5)
                summary["sample_data"] = sample_data
            
            return QueryResult(
                success=True,
                message=f"Sheet summary for '{sheet_name}'",
                data=[summary],
                row_count=1
            )
            
        except Exception as e:
            self.logger.error(f"Error getting sheet summary: {str(e)}")
            return QueryResult(
                success=False,
                message=f"Failed to get sheet summary: {str(e)}"
            )
    
    def find_records(self, sheet_name: str, search_term: str, columns: Optional[List[str]] = None) -> QueryResult:
        """
        Find records containing a search term in specified columns.
        
        Args:
            sheet_name: Name of the sheet to search
            search_term: Term to search for
            columns: Columns to search in (None for all columns)
            
        Returns:
            QueryResult: Matching records
        """
        try:
            # Get sheet and structure
            sheet = self.excel_service.get_sheet(sheet_name)
            if not sheet:
                return QueryResult(
                    success=False,
                    message=f"Sheet '{sheet_name}' not found"
                )
            
            structure = self.excel_service.get_structure()
            if not structure:
                return QueryResult(
                    success=False,
                    message="Unable to analyze sheet structure"
                )
            
            sheet_info = structure.get_sheet_info(sheet_name)
            headers = sheet_info.get('headers', [])
            
            # Extract data
            raw_data = self._extract_sheet_data(sheet, headers)
            
            # Search for matching records
            matching_records = []
            search_columns = columns if columns else headers
            
            for record in raw_data:
                for col in search_columns:
                    if col in record:
                        value = str(record[col]).lower()
                        if search_term.lower() in value:
                            matching_records.append(record)
                            break  # Found match in this record, move to next
            
            return QueryResult(
                success=True,
                message=f"Found {len(matching_records)} records containing '{search_term}'",
                data=matching_records,
                headers=headers,
                row_count=len(matching_records)
            )
            
        except Exception as e:
            self.logger.error(f"Error finding records: {str(e)}")
            return QueryResult(
                success=False,
                message=f"Failed to find records: {str(e)}"
            )
    
    def get_cross_sheet_data(self, sheet_references: List[Tuple[str, str]]) -> QueryResult:
        """
        Get data from multiple sheets with cross-sheet references.
        
        Args:
            sheet_references: List of (sheet_name, column_name) tuples
            
        Returns:
            QueryResult: Combined data from multiple sheets
        """
        try:
            combined_data = {}
            
            for sheet_name, column_name in sheet_references:
                # Get sheet data
                query = QueryData(
                    target_sheet=sheet_name,
                    columns=[column_name] if column_name != "*" else None
                )
                
                result = self.query_data(query)
                if not result.success:
                    return QueryResult(
                        success=False,
                        message=f"Failed to get data from sheet '{sheet_name}': {result.message}"
                    )
                
                # Store data with sheet prefix
                key = f"{sheet_name}.{column_name}" if column_name != "*" else sheet_name
                combined_data[key] = result.data
            
            return QueryResult(
                success=True,
                message=f"Retrieved cross-sheet data from {len(sheet_references)} references",
                data=[combined_data],
                row_count=1
            )
            
        except Exception as e:
            self.logger.error(f"Error getting cross-sheet data: {str(e)}")
            return QueryResult(
                success=False,
                message=f"Failed to get cross-sheet data: {str(e)}"
            )
    
    def _validate_query(self, query: QueryData) -> OperationResult:
        """
        Validate query parameters.
        
        Args:
            query: QueryData to validate
            
        Returns:
            OperationResult: Validation result
        """
        if not query.target_sheet:
            return OperationResult(
                success=False,
                message="Target sheet not specified"
            )
        
        # Check if sheet exists
        if query.target_sheet not in self.excel_service.get_sheet_names():
            return OperationResult(
                success=False,
                message=f"Sheet '{query.target_sheet}' does not exist"
            )
        
        # Validate sort order
        if query.sort_order not in ["asc", "desc"]:
            return OperationResult(
                success=False,
                message="Sort order must be 'asc' or 'desc'"
            )
        
        # Validate limit
        if query.limit is not None and query.limit < 1:
            return OperationResult(
                success=False,
                message="Limit must be greater than 0"
            )
        
        # Safety check: limit result size
        if query.limit is None or query.limit > 10000:
            return OperationResult(
                success=False,
                message="Query result limited to 10000 rows for safety. Please add a limit parameter."
            )
        
        return OperationResult(success=True, message="Validation passed")
    
    def _extract_sheet_data(self, sheet: Worksheet, headers: List[str], max_rows: Optional[int] = None) -> List[Dict[str, Any]]:
        """
        Extract data from worksheet into list of dictionaries.
        
        Args:
            sheet: Worksheet to extract from
            headers: Column headers
            max_rows: Maximum number of rows to extract
            
        Returns:
            List[Dict[str, Any]]: Extracted data
        """
        data = []
        
        # Determine row range
        start_row = 2  # Skip header row
        end_row = sheet.max_row + 1
        
        if max_rows:
            end_row = min(start_row + max_rows, end_row)
        
        for row_num in range(start_row, end_row):
            record = {}
            has_data = False
            
            for col_idx, header in enumerate(headers, 1):
                cell_value = sheet.cell(row=row_num, column=col_idx).value
                record[header] = cell_value
                
                if cell_value is not None:
                    has_data = True
            
            # Only add records that have at least one non-empty cell
            if has_data:
                data.append(record)
        
        return data
    
    def _apply_filters(self, data: List[Dict[str, Any]], conditions: Optional[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """
        Apply filter conditions to data.
        
        Args:
            data: Data to filter
            conditions: Filter conditions
            
        Returns:
            List[Dict[str, Any]]: Filtered data
        """
        if not conditions:
            return data
        
        filtered_data = []
        
        for record in data:
            matches = True
            
            for column, condition in conditions.items():
                if column not in record:
                    matches = False
                    break
                
                record_value = record[column]
                
                # Handle different condition types
                if isinstance(condition, dict):
                    # Complex condition like {"operator": ">=", "value": 100}
                    operator = condition.get("operator", "=")
                    value = condition.get("value")
                    
                    if not self._evaluate_condition(record_value, operator, value):
                        matches = False
                        break
                else:
                    # Simple equality condition
                    if record_value != condition:
                        matches = False
                        break
            
            if matches:
                filtered_data.append(record)
        
        return filtered_data
    
    def _evaluate_condition(self, record_value: Any, operator: str, condition_value: Any) -> bool:
        """
        Evaluate a single condition.
        
        Args:
            record_value: Value from the record
            operator: Comparison operator
            condition_value: Value to compare against
            
        Returns:
            bool: True if condition matches
        """
        try:
            if operator == "=":
                return record_value == condition_value
            elif operator == "!=":
                return record_value != condition_value
            elif operator == ">":
                return record_value > condition_value
            elif operator == ">=":
                return record_value >= condition_value
            elif operator == "<":
                return record_value < condition_value
            elif operator == "<=":
                return record_value <= condition_value
            elif operator == "contains":
                return str(condition_value).lower() in str(record_value).lower()
            elif operator == "starts_with":
                return str(record_value).lower().startswith(str(condition_value).lower())
            elif operator == "ends_with":
                return str(record_value).lower().endswith(str(condition_value).lower())
            else:
                return False
        except (TypeError, ValueError):
            return False
    
    def _select_columns(self, data: List[Dict[str, Any]], columns: Optional[List[str]], all_headers: List[str]) -> List[Dict[str, Any]]:
        """
        Select specific columns from data.
        
        Args:
            data: Data to select from
            columns: Columns to select (None for all)
            all_headers: All available headers
            
        Returns:
            List[Dict[str, Any]]: Data with selected columns
        """
        if not columns:
            return data
        
        selected_data = []
        
        for record in data:
            selected_record = {}
            for column in columns:
                if column in record:
                    selected_record[column] = record[column]
            selected_data.append(selected_record)
        
        return selected_data
    
    def _apply_sorting(self, data: List[Dict[str, Any]], sort_by: Optional[str], sort_order: str) -> List[Dict[str, Any]]:
        """
        Apply sorting to data.
        
        Args:
            data: Data to sort
            sort_by: Column to sort by
            sort_order: Sort order ("asc" or "desc")
            
        Returns:
            List[Dict[str, Any]]: Sorted data
        """
        if not sort_by or not data:
            return data
        
        try:
            reverse = sort_order == "desc"
            
            # Sort with None values handled
            def sort_key(record):
                value = record.get(sort_by)
                if value is None:
                    return (1, "")  # Put None values at the end
                return (0, value)
            
            return sorted(data, key=sort_key, reverse=reverse)
        except Exception as e:
            self.logger.warning(f"Failed to sort data: {str(e)}")
            return data
    
    def _apply_limit(self, data: List[Dict[str, Any]], limit: Optional[int]) -> List[Dict[str, Any]]:
        """
        Apply limit to data.
        
        Args:
            data: Data to limit
            limit: Maximum number of records
            
        Returns:
            List[Dict[str, Any]]: Limited data
        """
        if limit is None:
            return data
        
        return data[:limit]
    
    def _calculate_aggregations(self, data: List[Dict[str, Any]], aggregations: Dict[str, str]) -> Dict[str, Any]:
        """
        Calculate aggregation functions on data.
        
        Args:
            data: Data to aggregate
            aggregations: Column -> function mapping
            
        Returns:
            Dict[str, Any]: Aggregation results
        """
        results = {}
        
        for column, function in aggregations.items():
            try:
                # Extract column values (excluding None)
                values = [record.get(column) for record in data if record.get(column) is not None]
                
                if not values:
                    results[f"{function}({column})"] = None
                    continue
                
                # Calculate aggregation
                if function.lower() == "count":
                    results[f"{function}({column})"] = len(values)
                elif function.lower() == "sum":
                    numeric_values = [v for v in values if isinstance(v, (int, float))]
                    results[f"{function}({column})"] = sum(numeric_values) if numeric_values else 0
                elif function.lower() == "avg":
                    numeric_values = [v for v in values if isinstance(v, (int, float))]
                    results[f"{function}({column})"] = sum(numeric_values) / len(numeric_values) if numeric_values else 0
                elif function.lower() == "min":
                    results[f"{function}({column})"] = min(values)
                elif function.lower() == "max":
                    results[f"{function}({column})"] = max(values)
                else:
                    results[f"{function}({column})"] = f"Unknown function: {function}"
                    
            except Exception as e:
                results[f"{function}({column})"] = f"Error: {str(e)}"
        
        return results


@dataclass
class UpdateData:
    """Data structure for update operations."""
    target_sheet: str
    updates: Dict[str, Any]  # Column -> new value mapping
    conditions: Optional[Dict[str, Any]] = None  # Filter conditions to identify rows
    target_row: Optional[int] = None  # Specific row to update
    target_range: Optional[str] = None  # Specific range to update (e.g., "A1:C3")
    unique_identifier: Optional[Dict[str, Any]] = None  # Unique identifier for ambiguous updates


@dataclass
class UpdateResult:
    """Result of an update operation."""
    success: bool
    message: str
    affected_rows: int = 0
    affected_cells: int = 0
    changes_made: Optional[List[Dict[str, Any]]] = None
    warnings: List[str] = None
    
    def __post_init__(self):
        if self.warnings is None:
            self.warnings = []
        if self.changes_made is None:
            self.changes_made = []


class DataUpdateHandler:
    """Handler for data update operations (Update operations)."""
    
    def __init__(self, excel_service: ExcelService, safety_manager: SafetyManager):
        """
        Initialize the data update handler.
        
        Args:
            excel_service: Excel service instance
            safety_manager: Safety manager for validation
        """
        self.excel_service = excel_service
        self.safety_manager = safety_manager
        self.logger = logging.getLogger(__name__)
    
    def update_data(self, update_data: UpdateData) -> UpdateResult:
        """
        Update data in the Excel sheet based on conditions or specific targets.
        
        Args:
            update_data: UpdateData containing update parameters
            
        Returns:
            UpdateResult: Result of the update operation
        """
        try:
            # Validate update data
            validation_result = self._validate_update_data(update_data)
            if not validation_result.success:
                return UpdateResult(
                    success=False,
                    message=validation_result.message
                )
            
            # Get worksheet
            sheet = self.excel_service.get_sheet(update_data.target_sheet)
            if not sheet:
                return UpdateResult(
                    success=False,
                    message=f"Sheet '{update_data.target_sheet}' not found"
                )
            
            # Create backup before operation
            backup_path = self.excel_service.create_backup()
            if not backup_path:
                self.logger.warning("Failed to create backup before update operation")
            
            # Determine update strategy
            if update_data.target_row is not None:
                return self._update_specific_row(sheet, update_data)
            elif update_data.target_range is not None:
                return self._update_specific_range(sheet, update_data)
            elif update_data.conditions is not None:
                return self._update_by_conditions(sheet, update_data)
            elif update_data.unique_identifier is not None:
                return self._update_by_unique_identifier(sheet, update_data)
            else:
                return UpdateResult(
                    success=False,
                    message="No valid update target specified (row, range, conditions, or unique identifier)"
                )
            
        except Exception as e:
            self.logger.error(f"Error updating data: {str(e)}")
            return UpdateResult(
                success=False,
                message=f"Failed to update data: {str(e)}"
            )
    
    def update_cell(self, sheet_name: str, cell_reference: str, new_value: Any) -> UpdateResult:
        """
        Update a single cell value.
        
        Args:
            sheet_name: Name of the target sheet
            cell_reference: Cell reference (e.g., "A1", "B5")
            new_value: New value for the cell
            
        Returns:
            UpdateResult: Result of the update operation
        """
        try:
            # Get worksheet
            sheet = self.excel_service.get_sheet(sheet_name)
            if not sheet:
                return UpdateResult(
                    success=False,
                    message=f"Sheet '{sheet_name}' not found"
                )
            
            # Create backup before operation
            backup_path = self.excel_service.create_backup()
            if not backup_path:
                self.logger.warning("Failed to create backup before cell update")
            
            # Get current value for change tracking
            try:
                cell = sheet[cell_reference]
                old_value = cell.value
            except Exception:
                return UpdateResult(
                    success=False,
                    message=f"Invalid cell reference: {cell_reference}"
                )
            
            # Validate and convert new value
            converted_value = self._validate_and_convert_cell_value(
                new_value, sheet_name, cell_reference
            )
            
            # Update cell
            cell.value = converted_value
            
            # Save workbook
            if not self.excel_service.save_workbook(create_backup=False):
                return UpdateResult(
                    success=False,
                    message="Failed to save workbook after cell update"
                )
            
            # Update structure analysis
            self.excel_service.structure = self.excel_service._analyze_structure()
            
            change_record = {
                "cell": cell_reference,
                "old_value": old_value,
                "new_value": converted_value
            }
            
            return UpdateResult(
                success=True,
                message=f"Successfully updated cell {cell_reference}",
                affected_cells=1,
                changes_made=[change_record]
            )
            
        except Exception as e:
            self.logger.error(f"Error updating cell: {str(e)}")
            return UpdateResult(
                success=False,
                message=f"Failed to update cell: {str(e)}"
            )
    
    def update_range(self, sheet_name: str, range_reference: str, values: List[List[Any]]) -> UpdateResult:
        """
        Update a range of cells with new values.
        
        Args:
            sheet_name: Name of the target sheet
            range_reference: Range reference (e.g., "A1:C3")
            values: 2D list of values to update
            
        Returns:
            UpdateResult: Result of the update operation
        """
        try:
            # Get worksheet
            sheet = self.excel_service.get_sheet(sheet_name)
            if not sheet:
                return UpdateResult(
                    success=False,
                    message=f"Sheet '{sheet_name}' not found"
                )
            
            # Create backup before operation
            backup_path = self.excel_service.create_backup()
            if not backup_path:
                self.logger.warning("Failed to create backup before range update")
            
            # Get range
            try:
                cell_range = sheet[range_reference]
            except Exception:
                return UpdateResult(
                    success=False,
                    message=f"Invalid range reference: {range_reference}"
                )
            
            changes_made = []
            cells_updated = 0
            
            # Handle single cell range
            if hasattr(cell_range, 'value'):
                if values and len(values) > 0 and len(values[0]) > 0:
                    old_value = cell_range.value
                    new_value = values[0][0]
                    
                    converted_value = self._validate_and_convert_cell_value(
                        new_value, sheet_name, range_reference
                    )
                    
                    cell_range.value = converted_value
                    cells_updated = 1
                    
                    changes_made.append({
                        "cell": range_reference,
                        "old_value": old_value,
                        "new_value": converted_value
                    })
            else:
                # Handle multi-cell range
                for row_idx, row in enumerate(cell_range):
                    if row_idx >= len(values):
                        break
                    
                    if isinstance(row, tuple):
                        # Multiple cells in row
                        for col_idx, cell in enumerate(row):
                            if col_idx >= len(values[row_idx]):
                                break
                            
                            old_value = cell.value
                            new_value = values[row_idx][col_idx]
                            
                            converted_value = self._validate_and_convert_cell_value(
                                new_value, sheet_name, cell.coordinate
                            )
                            
                            cell.value = converted_value
                            cells_updated += 1
                            
                            changes_made.append({
                                "cell": cell.coordinate,
                                "old_value": old_value,
                                "new_value": converted_value
                            })
                    else:
                        # Single cell in row
                        if len(values[row_idx]) > 0:
                            old_value = row.value
                            new_value = values[row_idx][0]
                            
                            converted_value = self._validate_and_convert_cell_value(
                                new_value, sheet_name, row.coordinate
                            )
                            
                            row.value = converted_value
                            cells_updated += 1
                            
                            changes_made.append({
                                "cell": row.coordinate,
                                "old_value": old_value,
                                "new_value": converted_value
                            })
            
            # Save workbook
            if not self.excel_service.save_workbook(create_backup=False):
                return UpdateResult(
                    success=False,
                    message="Failed to save workbook after range update"
                )
            
            # Update structure analysis
            self.excel_service.structure = self.excel_service._analyze_structure()
            
            return UpdateResult(
                success=True,
                message=f"Successfully updated range {range_reference} with {cells_updated} cells",
                affected_cells=cells_updated,
                changes_made=changes_made
            )
            
        except Exception as e:
            self.logger.error(f"Error updating range: {str(e)}")
            return UpdateResult(
                success=False,
                message=f"Failed to update range: {str(e)}"
            )
    
    def _update_specific_row(self, sheet: Worksheet, update_data: UpdateData) -> UpdateResult:
        """
        Update a specific row by row number.
        
        Args:
            sheet: Target worksheet
            update_data: Update data containing row and values
            
        Returns:
            UpdateResult: Result of the update operation
        """
        try:
            target_row = update_data.target_row
            
            # Validate row number
            if target_row < 1 or target_row > sheet.max_row:
                return UpdateResult(
                    success=False,
                    message=f"Row {target_row} is out of range (1-{sheet.max_row})"
                )
            
            # Get sheet structure for column mapping
            structure = self.excel_service.get_structure()
            if not structure:
                return UpdateResult(
                    success=False,
                    message="Unable to analyze sheet structure"
                )
            
            sheet_info = structure.get_sheet_info(update_data.target_sheet)
            headers = sheet_info.get('headers', [])
            
            changes_made = []
            cells_updated = 0
            
            # Update each specified column
            for column_name, new_value in update_data.updates.items():
                if column_name not in headers:
                    continue
                
                col_idx = headers.index(column_name) + 1
                cell = sheet.cell(row=target_row, column=col_idx)
                old_value = cell.value
                
                # Validate and convert value
                converted_value = self._validate_and_convert_cell_value(
                    new_value, update_data.target_sheet, cell.coordinate
                )
                
                cell.value = converted_value
                cells_updated += 1
                
                changes_made.append({
                    "cell": cell.coordinate,
                    "column": column_name,
                    "old_value": old_value,
                    "new_value": converted_value
                })
            
            # Save workbook
            if not self.excel_service.save_workbook(create_backup=False):
                return UpdateResult(
                    success=False,
                    message="Failed to save workbook after row update"
                )
            
            # Update structure analysis
            self.excel_service.structure = self.excel_service._analyze_structure()
            
            return UpdateResult(
                success=True,
                message=f"Successfully updated row {target_row} with {cells_updated} changes",
                affected_rows=1,
                affected_cells=cells_updated,
                changes_made=changes_made
            )
            
        except Exception as e:
            self.logger.error(f"Error updating specific row: {str(e)}")
            return UpdateResult(
                success=False,
                message=f"Failed to update row: {str(e)}"
            )
    
    def _update_specific_range(self, sheet: Worksheet, update_data: UpdateData) -> UpdateResult:
        """
        Update a specific range of cells.
        
        Args:
            sheet: Target worksheet
            update_data: Update data containing range and values
            
        Returns:
            UpdateResult: Result of the update operation
        """
        try:
            # For range updates, we expect updates to contain a single "values" key
            # with a 2D list of values
            if "values" not in update_data.updates:
                return UpdateResult(
                    success=False,
                    message="Range update requires 'values' key with 2D list of values"
                )
            
            values = update_data.updates["values"]
            if not isinstance(values, list):
                return UpdateResult(
                    success=False,
                    message="Range update values must be a 2D list"
                )
            
            return self.update_range(
                update_data.target_sheet,
                update_data.target_range,
                values
            )
            
        except Exception as e:
            self.logger.error(f"Error updating specific range: {str(e)}")
            return UpdateResult(
                success=False,
                message=f"Failed to update range: {str(e)}"
            )
    
    def _update_by_conditions(self, sheet: Worksheet, update_data: UpdateData) -> UpdateResult:
        """
        Update rows that match specified conditions.
        
        Args:
            sheet: Target worksheet
            update_data: Update data containing conditions and values
            
        Returns:
            UpdateResult: Result of the update operation
        """
        try:
            # Get sheet structure
            structure = self.excel_service.get_structure()
            if not structure:
                return UpdateResult(
                    success=False,
                    message="Unable to analyze sheet structure"
                )
            
            sheet_info = structure.get_sheet_info(update_data.target_sheet)
            headers = sheet_info.get('headers', [])
            
            # Extract all data to find matching rows
            query_handler = DataQueryHandler(self.excel_service, self.safety_manager)
            raw_data = query_handler._extract_sheet_data(sheet, headers)
            
            # Find matching rows
            matching_rows = []
            for row_idx, record in enumerate(raw_data, 2):  # Start from row 2 (skip header)
                matches = True
                
                for column, condition in update_data.conditions.items():
                    if column not in record:
                        matches = False
                        break
                    
                    record_value = record[column]
                    
                    # Handle different condition types
                    if isinstance(condition, dict):
                        operator = condition.get("operator", "=")
                        value = condition.get("value")
                        
                        if not query_handler._evaluate_condition(record_value, operator, value):
                            matches = False
                            break
                    else:
                        if record_value != condition:
                            matches = False
                            break
                
                if matches:
                    matching_rows.append(row_idx)
            
            if not matching_rows:
                return UpdateResult(
                    success=True,
                    message="No rows matched the specified conditions",
                    affected_rows=0
                )
            
            # Safety check: limit number of rows that can be updated
            if len(matching_rows) > 50:
                return UpdateResult(
                    success=False,
                    message=f"Update would affect {len(matching_rows)} rows. Limited to 50 rows for safety."
                )
            
            # Update matching rows
            changes_made = []
            total_cells_updated = 0
            
            for row_num in matching_rows:
                for column_name, new_value in update_data.updates.items():
                    if column_name not in headers:
                        continue
                    
                    col_idx = headers.index(column_name) + 1
                    cell = sheet.cell(row=row_num, column=col_idx)
                    old_value = cell.value
                    
                    # Validate and convert value
                    converted_value = self._validate_and_convert_cell_value(
                        new_value, update_data.target_sheet, cell.coordinate
                    )
                    
                    cell.value = converted_value
                    total_cells_updated += 1
                    
                    changes_made.append({
                        "cell": cell.coordinate,
                        "row": row_num,
                        "column": column_name,
                        "old_value": old_value,
                        "new_value": converted_value
                    })
            
            # Save workbook
            if not self.excel_service.save_workbook(create_backup=False):
                return UpdateResult(
                    success=False,
                    message="Failed to save workbook after conditional update"
                )
            
            # Update structure analysis
            self.excel_service.structure = self.excel_service._analyze_structure()
            
            return UpdateResult(
                success=True,
                message=f"Successfully updated {len(matching_rows)} rows with {total_cells_updated} changes",
                affected_rows=len(matching_rows),
                affected_cells=total_cells_updated,
                changes_made=changes_made
            )
            
        except Exception as e:
            self.logger.error(f"Error updating by conditions: {str(e)}")
            return UpdateResult(
                success=False,
                message=f"Failed to update by conditions: {str(e)}"
            )
    
    def _update_by_unique_identifier(self, sheet: Worksheet, update_data: UpdateData) -> UpdateResult:
        """
        Update rows using unique identifier to resolve ambiguity.
        
        Args:
            sheet: Target worksheet
            update_data: Update data containing unique identifier and values
            
        Returns:
            UpdateResult: Result of the update operation
        """
        try:
            # Use unique identifier as conditions for finding the row
            temp_update_data = UpdateData(
                target_sheet=update_data.target_sheet,
                updates=update_data.updates,
                conditions=update_data.unique_identifier
            )
            
            result = self._update_by_conditions(sheet, temp_update_data)
            
            # Check if exactly one row was updated (unique identifier should be unique)
            if result.success and result.affected_rows > 1:
                result.warnings.append(
                    f"Unique identifier matched {result.affected_rows} rows. "
                    "Consider using a more specific identifier."
                )
            
            return result
            
        except Exception as e:
            self.logger.error(f"Error updating by unique identifier: {str(e)}")
            return UpdateResult(
                success=False,
                message=f"Failed to update by unique identifier: {str(e)}"
            )
    
    def _validate_update_data(self, update_data: UpdateData) -> OperationResult:
        """
        Validate update data and parameters.
        
        Args:
            update_data: UpdateData to validate
            
        Returns:
            OperationResult: Validation result
        """
        if not update_data.target_sheet:
            return OperationResult(
                success=False,
                message="Target sheet not specified"
            )
        
        # Check if sheet exists
        if update_data.target_sheet not in self.excel_service.get_sheet_names():
            return OperationResult(
                success=False,
                message=f"Sheet '{update_data.target_sheet}' does not exist"
            )
        
        if not update_data.updates:
            return OperationResult(
                success=False,
                message="No updates specified"
            )
        
        # Validate that at least one target is specified
        targets_specified = sum([
            update_data.target_row is not None,
            update_data.target_range is not None,
            update_data.conditions is not None,
            update_data.unique_identifier is not None
        ])
        
        if targets_specified == 0:
            return OperationResult(
                success=False,
                message="No update target specified (row, range, conditions, or unique identifier)"
            )
        
        if targets_specified > 1:
            return OperationResult(
                success=False,
                message="Multiple update targets specified. Please specify only one."
            )
        
        # Validate row number if specified
        if update_data.target_row is not None and update_data.target_row < 1:
            return OperationResult(
                success=False,
                message="Row number must be greater than 0"
            )
        
        return OperationResult(success=True, message="Validation passed")
    
    def _validate_and_convert_cell_value(self, value: Any, sheet_name: str, cell_reference: str) -> Any:
        """
        Validate and convert a cell value based on expected data type.
        
        Args:
            value: Value to validate and convert
            sheet_name: Name of the target sheet
            cell_reference: Cell reference for context
            
        Returns:
            Any: Converted value
        """
        if value is None or value == "":
            return None
        
        # For now, return the value as-is with basic type conversion
        # This could be enhanced to use column data type analysis
        try:
            # Try to preserve numeric types
            if isinstance(value, str):
                # Try to convert to number if it looks like one
                if value.replace('.', '').replace('-', '').isdigit():
                    if '.' in value:
                        return float(value)
                    else:
                        return int(value)
                # Try to convert to boolean
                elif value.lower() in ('true', 'false'):
                    return value.lower() == 'true'
            
            return value
        except (ValueError, TypeError):
            return str(value)


@dataclass
class DeletionData:
    """Data structure for deletion operations."""
    target_sheet: str
    conditions: Optional[Dict[str, Any]] = None  # Filter conditions to identify rows
    target_rows: Optional[List[int]] = None  # Specific rows to delete
    target_range: Optional[str] = None  # Specific range to delete
    unique_identifier: Optional[Dict[str, Any]] = None  # Unique identifier for ambiguous deletions
    confirmation_required: bool = True  # Whether to require confirmation


@dataclass
class DeletionResult:
    """Result of a deletion operation."""
    success: bool
    message: str
    affected_rows: int = 0
    deleted_data: Optional[List[Dict[str, Any]]] = None  # Data that was deleted (for recovery)
    warnings: List[str] = None
    requires_confirmation: bool = False
    confirmation_prompt: Optional[str] = None
    
    def __post_init__(self):
        if self.warnings is None:
            self.warnings = []
        if self.deleted_data is None:
            self.deleted_data = []


class DataDeletionHandler:
    """Handler for data deletion operations (Delete operations)."""
    
    def __init__(self, excel_service: ExcelService, safety_manager: SafetyManager):
        """
        Initialize the data deletion handler.
        
        Args:
            excel_service: Excel service instance
            safety_manager: Safety manager for validation
        """
        self.excel_service = excel_service
        self.safety_manager = safety_manager
        self.logger = logging.getLogger(__name__)
    
    def delete_data(self, deletion_data: DeletionData, confirmed: bool = False) -> DeletionResult:
        """
        Delete data from the Excel sheet based on conditions or specific targets.
        
        Args:
            deletion_data: DeletionData containing deletion parameters
            confirmed: Whether the user has confirmed the deletion
            
        Returns:
            DeletionResult: Result of the deletion operation
        """
        try:
            # Validate deletion data
            validation_result = self._validate_deletion_data(deletion_data)
            if not validation_result.success:
                return DeletionResult(
                    success=False,
                    message=validation_result.message
                )
            
            # Get worksheet
            sheet = self.excel_service.get_sheet(deletion_data.target_sheet)
            if not sheet:
                return DeletionResult(
                    success=False,
                    message=f"Sheet '{deletion_data.target_sheet}' not found"
                )
            
            # Determine what will be deleted for confirmation
            preview_result = self._preview_deletion(sheet, deletion_data)
            if not preview_result.success:
                return preview_result
            
            # Safety check: limit number of rows that can be deleted
            if preview_result.affected_rows > 50:
                return DeletionResult(
                    success=False,
                    message=f"Deletion would affect {preview_result.affected_rows} rows. "
                           f"Limited to 50 rows for safety. Please use more specific conditions."
                )
            
            # Check if confirmation is required and not yet provided
            if deletion_data.confirmation_required and not confirmed:
                return DeletionResult(
                    success=False,
                    message="Confirmation required for deletion operation",
                    requires_confirmation=True,
                    confirmation_prompt=self._generate_confirmation_prompt(preview_result),
                    affected_rows=preview_result.affected_rows,
                    deleted_data=preview_result.deleted_data
                )
            
            # Create backup before operation
            backup_path = self.excel_service.create_backup()
            if not backup_path:
                self.logger.warning("Failed to create backup before deletion operation")
            
            # Perform the actual deletion
            if deletion_data.target_rows is not None:
                return self._delete_specific_rows(sheet, deletion_data)
            elif deletion_data.target_range is not None:
                return self._delete_specific_range(sheet, deletion_data)
            elif deletion_data.conditions is not None:
                return self._delete_by_conditions(sheet, deletion_data)
            elif deletion_data.unique_identifier is not None:
                return self._delete_by_unique_identifier(sheet, deletion_data)
            else:
                return DeletionResult(
                    success=False,
                    message="No valid deletion target specified"
                )
            
        except Exception as e:
            self.logger.error(f"Error deleting data: {str(e)}")
            return DeletionResult(
                success=False,
                message=f"Failed to delete data: {str(e)}"
            )
    
    def delete_rows(self, sheet_name: str, row_numbers: List[int], confirmed: bool = False) -> DeletionResult:
        """
        Delete specific rows by row numbers.
        
        Args:
            sheet_name: Name of the target sheet
            row_numbers: List of row numbers to delete
            confirmed: Whether the user has confirmed the deletion
            
        Returns:
            DeletionResult: Result of the deletion operation
        """
        deletion_data = DeletionData(
            target_sheet=sheet_name,
            target_rows=row_numbers,
            confirmation_required=True
        )
        
        return self.delete_data(deletion_data, confirmed)
    
    def delete_by_condition(self, sheet_name: str, conditions: Dict[str, Any], confirmed: bool = False) -> DeletionResult:
        """
        Delete rows that match specified conditions.
        
        Args:
            sheet_name: Name of the target sheet
            conditions: Conditions to match for deletion
            confirmed: Whether the user has confirmed the deletion
            
        Returns:
            DeletionResult: Result of the deletion operation
        """
        deletion_data = DeletionData(
            target_sheet=sheet_name,
            conditions=conditions,
            confirmation_required=True
        )
        
        return self.delete_data(deletion_data, confirmed)
    
    def _preview_deletion(self, sheet: Worksheet, deletion_data: DeletionData) -> DeletionResult:
        """
        Preview what will be deleted without actually deleting.
        
        Args:
            sheet: Target worksheet
            deletion_data: Deletion data
            
        Returns:
            DeletionResult: Preview of what would be deleted
        """
        try:
            if deletion_data.target_rows is not None:
                return self._preview_specific_rows(sheet, deletion_data)
            elif deletion_data.target_range is not None:
                return self._preview_specific_range(sheet, deletion_data)
            elif deletion_data.conditions is not None:
                return self._preview_by_conditions(sheet, deletion_data)
            elif deletion_data.unique_identifier is not None:
                return self._preview_by_unique_identifier(sheet, deletion_data)
            else:
                return DeletionResult(
                    success=False,
                    message="No valid deletion target specified for preview"
                )
                
        except Exception as e:
            self.logger.error(f"Error previewing deletion: {str(e)}")
            return DeletionResult(
                success=False,
                message=f"Failed to preview deletion: {str(e)}"
            )
    
    def _preview_specific_rows(self, sheet: Worksheet, deletion_data: DeletionData) -> DeletionResult:
        """Preview deletion of specific rows."""
        try:
            row_numbers = deletion_data.target_rows
            
            # Get sheet structure
            structure = self.excel_service.get_structure()
            if not structure:
                return DeletionResult(
                    success=False,
                    message="Unable to analyze sheet structure"
                )
            
            sheet_info = structure.get_sheet_info(deletion_data.target_sheet)
            headers = sheet_info.get('headers', [])
            
            # Collect data that would be deleted
            deleted_data = []
            valid_rows = []
            
            for row_num in row_numbers:
                if row_num < 2 or row_num > sheet.max_row:  # Skip header row and invalid rows
                    continue
                
                record = {}
                for col_idx, header in enumerate(headers, 1):
                    cell_value = sheet.cell(row=row_num, column=col_idx).value
                    record[header] = cell_value
                
                deleted_data.append(record)
                valid_rows.append(row_num)
            
            return DeletionResult(
                success=True,
                message=f"Would delete {len(valid_rows)} rows",
                affected_rows=len(valid_rows),
                deleted_data=deleted_data
            )
            
        except Exception as e:
            return DeletionResult(
                success=False,
                message=f"Failed to preview row deletion: {str(e)}"
            )
    
    def _preview_specific_range(self, sheet: Worksheet, deletion_data: DeletionData) -> DeletionResult:
        """Preview deletion of specific range."""
        try:
            # For range deletion, we'll clear the range content
            # This is different from row deletion which removes entire rows
            try:
                cell_range = sheet[deletion_data.target_range]
            except Exception:
                return DeletionResult(
                    success=False,
                    message=f"Invalid range reference: {deletion_data.target_range}"
                )
            
            # Count cells that would be affected
            cells_to_clear = 0
            deleted_data = []
            
            if hasattr(cell_range, 'value'):
                # Single cell
                if cell_range.value is not None:
                    cells_to_clear = 1
                    deleted_data.append({
                        "cell": deletion_data.target_range,
                        "value": cell_range.value
                    })
            else:
                # Multiple cells
                for row in cell_range:
                    if isinstance(row, tuple):
                        for cell in row:
                            if cell.value is not None:
                                cells_to_clear += 1
                                deleted_data.append({
                                    "cell": cell.coordinate,
                                    "value": cell.value
                                })
                    else:
                        if row.value is not None:
                            cells_to_clear += 1
                            deleted_data.append({
                                "cell": row.coordinate,
                                "value": row.value
                            })
            
            return DeletionResult(
                success=True,
                message=f"Would clear {cells_to_clear} cells in range {deletion_data.target_range}",
                affected_rows=0,  # Range deletion doesn't remove rows
                deleted_data=deleted_data
            )
            
        except Exception as e:
            return DeletionResult(
                success=False,
                message=f"Failed to preview range deletion: {str(e)}"
            )
    
    def _preview_by_conditions(self, sheet: Worksheet, deletion_data: DeletionData) -> DeletionResult:
        """Preview deletion by conditions."""
        try:
            # Get sheet structure
            structure = self.excel_service.get_structure()
            if not structure:
                return DeletionResult(
                    success=False,
                    message="Unable to analyze sheet structure"
                )
            
            sheet_info = structure.get_sheet_info(deletion_data.target_sheet)
            headers = sheet_info.get('headers', [])
            
            # Use query handler to find matching rows
            query_handler = DataQueryHandler(self.excel_service, self.safety_manager)
            raw_data = query_handler._extract_sheet_data(sheet, headers)
            
            # Find matching rows
            matching_records = []
            for record in raw_data:
                matches = True
                
                for column, condition in deletion_data.conditions.items():
                    if column not in record:
                        matches = False
                        break
                    
                    record_value = record[column]
                    
                    # Handle different condition types
                    if isinstance(condition, dict):
                        operator = condition.get("operator", "=")
                        value = condition.get("value")
                        
                        if not query_handler._evaluate_condition(record_value, operator, value):
                            matches = False
                            break
                    else:
                        if record_value != condition:
                            matches = False
                            break
                
                if matches:
                    matching_records.append(record)
            
            return DeletionResult(
                success=True,
                message=f"Would delete {len(matching_records)} rows matching conditions",
                affected_rows=len(matching_records),
                deleted_data=matching_records
            )
            
        except Exception as e:
            return DeletionResult(
                success=False,
                message=f"Failed to preview conditional deletion: {str(e)}"
            )
    
    def _preview_by_unique_identifier(self, sheet: Worksheet, deletion_data: DeletionData) -> DeletionResult:
        """Preview deletion by unique identifier."""
        # Use unique identifier as conditions
        temp_deletion_data = DeletionData(
            target_sheet=deletion_data.target_sheet,
            conditions=deletion_data.unique_identifier
        )
        
        return self._preview_by_conditions(sheet, temp_deletion_data)
    
    def _delete_specific_rows(self, sheet: Worksheet, deletion_data: DeletionData) -> DeletionResult:
        """Delete specific rows by row numbers."""
        try:
            row_numbers = sorted(deletion_data.target_rows, reverse=True)  # Delete from bottom up
            
            # Get data before deletion for recovery
            preview_result = self._preview_specific_rows(sheet, deletion_data)
            if not preview_result.success:
                return preview_result
            
            deleted_count = 0
            for row_num in row_numbers:
                if row_num < 2 or row_num > sheet.max_row:  # Skip header row and invalid rows
                    continue
                
                sheet.delete_rows(row_num)
                deleted_count += 1
            
            # Save workbook
            if not self.excel_service.save_workbook(create_backup=False):
                return DeletionResult(
                    success=False,
                    message="Failed to save workbook after row deletion"
                )
            
            # Update structure analysis
            self.excel_service.structure = self.excel_service._analyze_structure()
            
            return DeletionResult(
                success=True,
                message=f"Successfully deleted {deleted_count} rows",
                affected_rows=deleted_count,
                deleted_data=preview_result.deleted_data
            )
            
        except Exception as e:
            self.logger.error(f"Error deleting specific rows: {str(e)}")
            return DeletionResult(
                success=False,
                message=f"Failed to delete rows: {str(e)}"
            )
    
    def _delete_specific_range(self, sheet: Worksheet, deletion_data: DeletionData) -> DeletionResult:
        """Delete (clear) specific range of cells."""
        try:
            # Get data before deletion for recovery
            preview_result = self._preview_specific_range(sheet, deletion_data)
            if not preview_result.success:
                return preview_result
            
            # Clear the range
            try:
                cell_range = sheet[deletion_data.target_range]
            except Exception:
                return DeletionResult(
                    success=False,
                    message=f"Invalid range reference: {deletion_data.target_range}"
                )
            
            cells_cleared = 0
            
            if hasattr(cell_range, 'value'):
                # Single cell
                if cell_range.value is not None:
                    cell_range.value = None
                    cells_cleared = 1
            else:
                # Multiple cells
                for row in cell_range:
                    if isinstance(row, tuple):
                        for cell in row:
                            if cell.value is not None:
                                cell.value = None
                                cells_cleared += 1
                    else:
                        if row.value is not None:
                            row.value = None
                            cells_cleared += 1
            
            # Save workbook
            if not self.excel_service.save_workbook(create_backup=False):
                return DeletionResult(
                    success=False,
                    message="Failed to save workbook after range deletion"
                )
            
            # Update structure analysis
            self.excel_service.structure = self.excel_service._analyze_structure()
            
            return DeletionResult(
                success=True,
                message=f"Successfully cleared {cells_cleared} cells in range {deletion_data.target_range}",
                affected_rows=0,
                deleted_data=preview_result.deleted_data
            )
            
        except Exception as e:
            self.logger.error(f"Error deleting specific range: {str(e)}")
            return DeletionResult(
                success=False,
                message=f"Failed to delete range: {str(e)}"
            )
    
    def _delete_by_conditions(self, sheet: Worksheet, deletion_data: DeletionData) -> DeletionResult:
        """Delete rows that match specified conditions."""
        try:
            # Get preview to find matching rows
            preview_result = self._preview_by_conditions(sheet, deletion_data)
            if not preview_result.success:
                return preview_result
            
            if preview_result.affected_rows == 0:
                return DeletionResult(
                    success=True,
                    message="No rows matched the specified conditions",
                    affected_rows=0
                )
            
            # Get sheet structure
            structure = self.excel_service.get_structure()
            if not structure:
                return DeletionResult(
                    success=False,
                    message="Unable to analyze sheet structure"
                )
            
            sheet_info = structure.get_sheet_info(deletion_data.target_sheet)
            headers = sheet_info.get('headers', [])
            
            # Find matching row numbers
            query_handler = DataQueryHandler(self.excel_service, self.safety_manager)
            raw_data = query_handler._extract_sheet_data(sheet, headers)
            
            matching_row_numbers = []
            for row_idx, record in enumerate(raw_data, 2):  # Start from row 2 (skip header)
                matches = True
                
                for column, condition in deletion_data.conditions.items():
                    if column not in record:
                        matches = False
                        break
                    
                    record_value = record[column]
                    
                    # Handle different condition types
                    if isinstance(condition, dict):
                        operator = condition.get("operator", "=")
                        value = condition.get("value")
                        
                        if not query_handler._evaluate_condition(record_value, operator, value):
                            matches = False
                            break
                    else:
                        if record_value != condition:
                            matches = False
                            break
                
                if matches:
                    matching_row_numbers.append(row_idx)
            
            # Delete rows (from bottom up to maintain row numbers)
            deleted_count = 0
            for row_num in sorted(matching_row_numbers, reverse=True):
                sheet.delete_rows(row_num)
                deleted_count += 1
            
            # Save workbook
            if not self.excel_service.save_workbook(create_backup=False):
                return DeletionResult(
                    success=False,
                    message="Failed to save workbook after conditional deletion"
                )
            
            # Update structure analysis
            self.excel_service.structure = self.excel_service._analyze_structure()
            
            return DeletionResult(
                success=True,
                message=f"Successfully deleted {deleted_count} rows matching conditions",
                affected_rows=deleted_count,
                deleted_data=preview_result.deleted_data
            )
            
        except Exception as e:
            self.logger.error(f"Error deleting by conditions: {str(e)}")
            return DeletionResult(
                success=False,
                message=f"Failed to delete by conditions: {str(e)}"
            )
    
    def _delete_by_unique_identifier(self, sheet: Worksheet, deletion_data: DeletionData) -> DeletionResult:
        """Delete rows using unique identifier."""
        # Use unique identifier as conditions
        temp_deletion_data = DeletionData(
            target_sheet=deletion_data.target_sheet,
            conditions=deletion_data.unique_identifier,
            confirmation_required=deletion_data.confirmation_required
        )
        
        result = self._delete_by_conditions(sheet, temp_deletion_data)
        
        # Check if exactly one row was deleted (unique identifier should be unique)
        if result.success and result.affected_rows > 1:
            result.warnings.append(
                f"Unique identifier matched {result.affected_rows} rows. "
                "Consider using a more specific identifier."
            )
        
        return result
    
    def _validate_deletion_data(self, deletion_data: DeletionData) -> OperationResult:
        """
        Validate deletion data and parameters.
        
        Args:
            deletion_data: DeletionData to validate
            
        Returns:
            OperationResult: Validation result
        """
        if not deletion_data.target_sheet:
            return OperationResult(
                success=False,
                message="Target sheet not specified"
            )
        
        # Check if sheet exists
        if deletion_data.target_sheet not in self.excel_service.get_sheet_names():
            return OperationResult(
                success=False,
                message=f"Sheet '{deletion_data.target_sheet}' does not exist"
            )
        
        # Validate that at least one target is specified
        targets_specified = sum([
            deletion_data.target_rows is not None,
            deletion_data.target_range is not None,
            deletion_data.conditions is not None,
            deletion_data.unique_identifier is not None
        ])
        
        if targets_specified == 0:
            return OperationResult(
                success=False,
                message="No deletion target specified (rows, range, conditions, or unique identifier)"
            )
        
        if targets_specified > 1:
            return OperationResult(
                success=False,
                message="Multiple deletion targets specified. Please specify only one."
            )
        
        # Validate row numbers if specified
        if deletion_data.target_rows is not None:
            for row_num in deletion_data.target_rows:
                if row_num < 2:  # Don't allow deleting header row
                    return OperationResult(
                        success=False,
                        message=f"Cannot delete header row or row {row_num}. Row numbers must be >= 2."
                    )
        
        return OperationResult(success=True, message="Validation passed")
    
    def _generate_confirmation_prompt(self, preview_result: DeletionResult) -> str:
        """
        Generate a confirmation prompt for the deletion operation.
        
        Args:
            preview_result: Preview result containing what would be deleted
            
        Returns:
            str: Confirmation prompt
        """
        if preview_result.affected_rows == 0:
            return "No data would be deleted. Proceed anyway?"
        
        prompt = f"This will delete {preview_result.affected_rows} row(s). "
        
        if preview_result.deleted_data and len(preview_result.deleted_data) > 0:
            # Show sample of data that would be deleted
            sample_size = min(3, len(preview_result.deleted_data))
            prompt += "Sample of data to be deleted:\n"
            
            for i, record in enumerate(preview_result.deleted_data[:sample_size]):
                if isinstance(record, dict):
                    # Row data
                    sample_values = []
                    for key, value in list(record.items())[:3]:  # Show first 3 columns
                        sample_values.append(f"{key}: {value}")
                    prompt += f"  Row {i+1}: {', '.join(sample_values)}\n"
                else:
                    # Cell data
                    prompt += f"  {record}\n"
            
            if len(preview_result.deleted_data) > sample_size:
                prompt += f"  ... and {len(preview_result.deleted_data) - sample_size} more\n"
        
        prompt += "\nAre you sure you want to proceed with this deletion? (This action cannot be undone without restoring from backup)"
        
        return prompt


# Global instances for use by template system
data_insertion_handler = DataInsertionHandler(None, None)  # Will be initialized with proper services
data_query_handler = DataQueryHandler(None, None)  # Will be initialized with proper services
data_update_handler = DataUpdateHandler(None, None)  # Will be initialized with proper services
data_deletion_handler = DataDeletionHandler(None, None)  # Will be initialized with proper services


# Wrapper functions for template system compatibility
def insert_row(sheet_name: str, data: dict, position: Optional[int] = None) -> Dict[str, Any]:
    """Wrapper function for row insertion."""
    return {
        "success": True,
        "message": f"Row insertion operation prepared for sheet '{sheet_name}' (placeholder implementation)",
        "affected_rows": 1,
        "operation": "insert_row",
        "parameters": {"sheet_name": sheet_name, "data": data, "position": position}
    }


def insert_column(sheet_name: str, column_name: str, values: list, position: Optional[str] = None) -> Dict[str, Any]:
    """Wrapper function for column insertion."""
    return {
        "success": True,
        "message": f"Column insertion operation prepared for sheet '{sheet_name}' (placeholder implementation)",
        "affected_rows": len(values) if values else 0,
        "operation": "insert_column",
        "parameters": {"sheet_name": sheet_name, "column_name": column_name, "values": values, "position": position}
    }


def update_cells(sheet_name: str, range_ref: str, values: Any, conditions: Optional[dict] = None) -> Dict[str, Any]:
    """Wrapper function for cell updates."""
    return {
        "success": True,
        "message": f"Cell update operation prepared for sheet '{sheet_name}' range '{range_ref}' (placeholder implementation)",
        "affected_rows": 1,
        "operation": "update_cells",
        "parameters": {"sheet_name": sheet_name, "range_ref": range_ref, "values": values, "conditions": conditions}
    }


def delete_rows(sheet_name: str, conditions: dict, max_rows: int = 50) -> Dict[str, Any]:
    """Wrapper function for row deletion."""
    return {
        "success": True,
        "message": f"Row deletion operation prepared for sheet '{sheet_name}' (placeholder implementation)",
        "affected_rows": 0,
        "operation": "delete_rows",
        "parameters": {"sheet_name": sheet_name, "conditions": conditions, "max_rows": max_rows}
    }

def query_data(sheet_name: str = None, columns: List[str] = None, conditions: dict = None, 
               limit: int = 100, sort_by: str = None, **kwargs) -> Dict[str, Any]:
    """
    Standalone function for querying Excel data.
    
    Args:
        sheet_name: Name of the sheet to query
        columns: List of columns to select (None for all)
        conditions: Filter conditions
        limit: Maximum number of rows to return
        sort_by: Column to sort by
        **kwargs: Additional parameters
        
    Returns:
        Dict with query results
    """
    try:
        # Import here to avoid circular imports
        from excel.excel_service import ExcelService
        from safety.safety_manager import SafetyManager
        
        # Get the global excel service instance (this is a simplified approach)
        # In a real implementation, this would be passed as a parameter
        import sys
        if hasattr(sys.modules.get('__main__'), 'excel_service'):
            excel_service = sys.modules['__main__'].excel_service
        else:
            # Fallback: try to get from current context
            excel_service = kwargs.get('excel_service')
            if not excel_service:
                return {
                    "success": False,
                    "message": "Excel service not available",
                    "data": None
                }
        
        # Create query handler
        safety_manager = SafetyManager()
        query_handler = DataQueryHandler(excel_service, safety_manager)
        
        # Determine sheet name
        if not sheet_name:
            # Use first available sheet
            sheet_names = excel_service.get_sheet_names()
            if not sheet_names:
                return {
                    "success": False,
                    "message": "No sheets available",
                    "data": None
                }
            sheet_name = sheet_names[0]
            print(f"Using default sheet: '{sheet_name}'")
        
        # Validate sheet exists
        available_sheets = excel_service.get_sheet_names()
        if sheet_name not in available_sheets:
            # Try to find a close match
            sheet_name_lower = sheet_name.lower()
            for available_sheet in available_sheets:
                if sheet_name_lower in available_sheet.lower() or available_sheet.lower() in sheet_name_lower:
                    sheet_name = available_sheet
                    print(f"Using matched sheet: '{sheet_name}'")
                    break
            else:
                return {
                    "success": False,
                    "message": f"Sheet '{sheet_name}' not found. Available sheets: {', '.join(available_sheets)}",
                    "data": None
                }
        
        # Handle special cases for "first" and "last" rows
        original_command = str(kwargs.get('original_command', '')).lower()
        import re  # Import re module at the beginning
        
        if 'first' in original_command:
            # Extract number from command like "first 2 rows"
            match = re.search(r'first\s+(\d+)', original_command)
            if match:
                limit = int(match.group(1))
                print(f"Extracting first {limit} rows")
        elif 'last' in original_command:
            # For "last N rows", we need to get all data and take the last N
            match = re.search(r'last\s+(\d+)', original_command)
            if match:
                requested_limit = int(match.group(1))
                print(f"Extracting last {requested_limit} rows")
                # Get all data first
                query_data_obj = QueryData(
                    target_sheet=sheet_name,
                    columns=columns,
                    conditions=conditions,
                    sort_by=sort_by,
                    limit=10000  # Get more data to find the last rows
                )
                result = query_handler.query_data(query_data_obj)
                if result.success and result.data:
                    # Take the last N rows
                    last_rows = result.data[-requested_limit:] if len(result.data) >= requested_limit else result.data
                    return {
                        "success": True,
                        "message": f"Retrieved last {len(last_rows)} rows from sheet '{sheet_name}'",
                        "data": last_rows,
                        "headers": result.headers,
                        "row_count": len(last_rows)
                    }
        
        # Normalize conditions parameter
        if isinstance(conditions, str):
            # If conditions is a string, try to interpret it as a search term
            # For now, we'll just ignore string conditions and use None
            print(f"Warning: conditions parameter is a string '{conditions}', ignoring for now")
            conditions = None
        elif conditions and not isinstance(conditions, dict):
            print(f"Warning: conditions parameter is not a dict: {type(conditions)}, ignoring")
            conditions = None
        
        # Create query object
        query_data_obj = QueryData(
            target_sheet=sheet_name,
            columns=columns,
            conditions=conditions,
            sort_by=sort_by,
            limit=limit
        )
        
        # Execute query
        result = query_handler.query_data(query_data_obj)
        
        # Convert QueryResult to dictionary format
        return {
            "success": result.success,
            "message": result.message,
            "data": result.data,
            "headers": result.headers,
            "row_count": result.row_count
        }
        
    except Exception as e:
        return {
            "success": False,
            "message": f"Error querying data: {str(e)}",
            "data": None
        }


def insert_row(sheet_name: str = None, data: List[Any] = None, column_names: List[str] = None,
               target_row: int = None, target_column: str = None, **kwargs) -> Dict[str, Any]:
    """
    Standalone function for inserting rows into Excel.
    
    Args:
        sheet_name: Name of the sheet
        data: List of values to insert
        column_names: Column names for the data
        target_row: Target row position
        target_column: Target column position
        **kwargs: Additional parameters
        
    Returns:
        Dict with insertion results
    """
    try:
        # Import here to avoid circular imports
        from excel.excel_service import ExcelService
        from safety.safety_manager import SafetyManager
        
        # Get the global excel service instance
        import sys
        if hasattr(sys.modules.get('__main__'), 'excel_service'):
            excel_service = sys.modules['__main__'].excel_service
        else:
            excel_service = kwargs.get('excel_service')
            if not excel_service:
                return {
                    "success": False,
                    "message": "Excel service not available",
                    "data": None
                }
        
        # Create insertion handler
        safety_manager = SafetyManager()
        insertion_handler = DataInsertionHandler(excel_service, safety_manager)
        
        # Determine sheet name
        if not sheet_name:
            sheet_names = excel_service.get_sheet_names()
            if not sheet_names:
                return {
                    "success": False,
                    "message": "No sheets available",
                    "data": None
                }
            sheet_name = sheet_names[0]
        
        # Create insertion data object
        insertion_data = InsertionData(
            values=data or [],
            target_sheet=sheet_name,
            target_row=target_row,
            target_column=target_column,
            column_names=column_names
        )
        
        # Execute insertion
        result = insertion_handler.insert_row(insertion_data)
        
        # Convert OperationResult to dictionary format
        return {
            "success": result.success,
            "message": result.message,
            "data": result.data,
            "affected_rows": result.affected_rows
        }
        
    except Exception as e:
        return {
            "success": False,
            "message": f"Error inserting row: {str(e)}",
            "data": None
        }


# Wrapper functions for template registry integration
def insert_row(excel_service: ExcelService, sheet_name: str = "", data: Any = None, 
               column_names: Optional[List[str]] = None, target_row: Optional[int] = None, 
               target_column: Optional[str] = None, **kwargs) -> Dict[str, Any]:
    """
    Wrapper function for insert_row operation that can be called by template registry.
    
    This function handles the parameter conversion from template registry format
    to the InsertionData format expected by DataInsertionHandler.
    
    Args:
        excel_service: Excel service instance
        sheet_name: Target sheet name
        data: Data to insert (can be dict, list, or single value)
        column_names: Optional column names
        target_row: Optional target row
        target_column: Optional target column
        **kwargs: Additional parameters
        
    Returns:
        Dict with operation result
    """
    from safety.safety_manager import SafetyManager
    
    # Create safety manager instance
    safety_manager = SafetyManager()
    
    # Create data insertion handler
    handler = DataInsertionHandler(excel_service, safety_manager)
    
    # Convert data parameter to list format
    if isinstance(data, dict):
        # Get sheet headers to determine column order
        if sheet_name and excel_service.workbook:
            try:
                sheet = excel_service.get_sheet(sheet_name)
                if sheet and sheet.max_row > 0:
                    # Get headers from first row
                    headers = []
                    for col in range(1, sheet.max_column + 1):
                        header = sheet.cell(row=1, column=col).value
                        if header:
                            headers.append(str(header))
                    
                    # Convert dict to list using header order
                    values_list = []
                    for header in headers:
                        if header in data:
                            values_list.append(data[header])
                        else:
                            values_list.append(None)  # Empty cell for missing data
                    
                    data_values = values_list
                else:
                    # No headers found, use dict values in order
                    data_values = list(data.values())
            except Exception as e:
                logging.getLogger(__name__).warning(f"Error getting sheet headers: {e}")
                # Fallback to dict values
                data_values = list(data.values())
        else:
            # No sheet specified or workbook not loaded, use dict values
            data_values = list(data.values())
    elif isinstance(data, list):
        data_values = data
    else:
        data_values = [data] if data is not None else []
    
    # Create InsertionData object
    insertion_data = InsertionData(
        values=data_values,
        target_sheet=sheet_name,
        target_row=target_row,
        target_column=target_column,
        column_names=column_names
    )
    
    # Execute the insertion
    result = handler.insert_row(insertion_data)
    
    # Return result in dictionary format
    return {
        'success': result.success,
        'message': result.message,
        'data': result.data,
        'affected_rows': result.affected_rows,
        'affected_columns': result.affected_columns,
        'warnings': result.warnings
    }


def query_data(excel_service: ExcelService, sheet_name: str = "", columns: Optional[List[str]] = None,
               conditions: Optional[Dict[str, Any]] = None, limit: Optional[int] = None,
               sort_by: Optional[str] = None, **kwargs) -> Dict[str, Any]:
    """
    Wrapper function for query_data operation that can be called by template registry.
    
    Args:
        excel_service: Excel service instance
        sheet_name: Target sheet name
        columns: Columns to select
        conditions: Filter conditions
        limit: Result limit
        sort_by: Sort column
        **kwargs: Additional parameters
        
    Returns:
        Dict with operation result
    """
    from safety.safety_manager import SafetyManager
    
    # Create safety manager instance
    safety_manager = SafetyManager()
    
    # Create data query handler
    handler = DataQueryHandler(excel_service, safety_manager)
    
    # Create QueryData object
    query_data_obj = QueryData(
        target_sheet=sheet_name,
        columns=columns,
        conditions=conditions,
        limit=limit or 100,
        sort_by=sort_by
    )
    
    # Execute the query
    result = handler.query_data(query_data_obj)
    
    # Return result in dictionary format
    return {
        'success': result.success,
        'message': result.message,
        'data': result.data,
        'headers': result.headers,
        'row_count': result.row_count
    }


def insert_column(excel_service: ExcelService, sheet_name: str = "", column_name: str = "",
                  values: Optional[List[Any]] = None, position: Optional[str] = None, **kwargs) -> Dict[str, Any]:
    """
    Wrapper function for insert_column operation that can be called by template registry.
    
    Args:
        excel_service: Excel service instance
        sheet_name: Target sheet name
        column_name: Name of the new column
        values: Values to insert in the column
        position: Position to insert the column
        **kwargs: Additional parameters
        
    Returns:
        Dict with operation result
    """
    from safety.safety_manager import SafetyManager
    
    # Create safety manager instance
    safety_manager = SafetyManager()
    
    # Create data insertion handler
    handler = DataInsertionHandler(excel_service, safety_manager)
    
    # Create InsertionData object for column insertion
    insertion_data = InsertionData(
        values=values or [],
        target_sheet=sheet_name,
        target_column=position,
        column_names=[column_name] if column_name else None,
        insert_type="column"
    )
    
    # Execute the insertion
    result = handler.insert_column(insertion_data)
    
    # Return result in dictionary format
    return {
        'success': result.success,
        'message': result.message,
        'data': result.data,
        'affected_rows': result.affected_rows,
        'affected_columns': result.affected_columns,
        'warnings': result.warnings
    }


def update_cells(excel_service: ExcelService, sheet_name: str = "", range: str = "",
                 values: Any = None, conditions: Optional[Dict[str, Any]] = None, **kwargs) -> Dict[str, Any]:
    """
    Wrapper function for update_cells operation that can be called by template registry.
    
    Args:
        excel_service: Excel service instance
        sheet_name: Target sheet name
        range: Cell range to update (e.g., "A1:B5") OR conditions dict for conditional updates
        values: Values to update (can be single value, list, or dict)
        conditions: Conditions for conditional updates
        **kwargs: Additional parameters
        
    Returns:
        Dict with operation result
    """
    from safety.safety_manager import SafetyManager
    
    # Create safety manager instance
    safety_manager = SafetyManager()
    
    try:
        if not excel_service.workbook:
            return {
                'success': False,
                'message': 'No Excel file loaded'
            }
        
        sheet = excel_service.get_sheet(sheet_name)
        if not sheet:
            return {
                'success': False,
                'message': f'Sheet "{sheet_name}" not found'
            }
        
        # Check if range is actually a conditions dict (from LLM parsing)
        actual_conditions = conditions
        if isinstance(range, dict):
            # Range parameter contains conditions, not an actual range
            actual_conditions = range
            range = None
        
        # Handle conditional updates (like "update row where Name=Lakshmi")
        if actual_conditions and isinstance(values, dict):
            return _update_rows_by_conditions(excel_service, sheet, sheet_name, actual_conditions, values)
        
        # Handle direct range updates (like "A1:B2")
        if range and values is not None:
            # Parse range (simple implementation for single cells like "A1")
            if ':' not in range:
                # Single cell update
                from openpyxl.utils.cell import coordinate_from_string
                try:
                    col, row = coordinate_from_string(range)
                    sheet[range].value = values
                    
                    # Save workbook
                    excel_service.save_workbook()
                    
                    return {
                        'success': True,
                        'message': f'Updated cell {range} with value: {values}',
                        'data': {'range': range, 'value': values},
                        'affected_rows': 1,
                        'affected_columns': 1
                    }
                except Exception as e:
                    return {
                        'success': False,
                        'message': f'Error updating cell {range}: {str(e)}'
                    }
        
        return {
            'success': False,
            'message': 'Update operation requires either a valid range or conditions with values'
        }
        
    except Exception as e:
        return {
            'success': False,
            'message': f'Error in update_cells: {str(e)}'
        }


def _update_rows_by_conditions(excel_service: ExcelService, sheet, sheet_name: str, 
                              conditions: Dict[str, Any], values: Dict[str, Any]) -> Dict[str, Any]:
    """
    Update rows that match the given conditions.
    
    Args:
        excel_service: Excel service instance
        sheet: Worksheet object
        sheet_name: Name of the sheet
        conditions: Dictionary of conditions to match (e.g., {"Name": "Lakshmi"})
        values: Dictionary of values to update (e.g., {"Active": True})
        
    Returns:
        Dict with operation result
    """
    try:
        # Get headers from first row
        headers = []
        for col in range(1, sheet.max_column + 1):
            header_cell = sheet.cell(row=1, column=col)
            if header_cell.value:
                headers.append(str(header_cell.value))
            else:
                break
        
        if not headers:
            return {
                'success': False,
                'message': 'No headers found in sheet'
            }
        
        # Create column index mapping
        header_to_col = {header: idx + 1 for idx, header in enumerate(headers)}
        
        # Validate condition columns exist
        for condition_col in conditions.keys():
            if condition_col not in header_to_col:
                return {
                    'success': False,
                    'message': f'Condition column "{condition_col}" not found in sheet headers: {headers}'
                }
        
        # Validate update columns exist
        for update_col in values.keys():
            if update_col not in header_to_col:
                return {
                    'success': False,
                    'message': f'Update column "{update_col}" not found in sheet headers: {headers}'
                }
        
        # Find and update matching rows
        updated_rows = 0
        
        for row_num in range(2, sheet.max_row + 1):  # Skip header row
            # Check if this row matches all conditions
            row_matches = True
            
            for condition_col, condition_value in conditions.items():
                col_idx = header_to_col[condition_col]
                cell_value = sheet.cell(row=row_num, column=col_idx).value
                
                # Convert both values to strings for comparison
                cell_str = str(cell_value) if cell_value is not None else ""
                condition_str = str(condition_value)
                
                if cell_str != condition_str:
                    row_matches = False
                    break
            
            # If row matches, update the specified columns
            if row_matches:
                for update_col, update_value in values.items():
                    col_idx = header_to_col[update_col]
                    sheet.cell(row=row_num, column=col_idx).value = update_value
                
                updated_rows += 1
        
        if updated_rows > 0:
            # Save workbook
            excel_service.save_workbook()
            
            return {
                'success': True,
                'message': f'Successfully updated {updated_rows} row(s) matching conditions {conditions}',
                'data': {
                    'conditions': conditions,
                    'updates': values,
                    'rows_updated': updated_rows
                },
                'affected_rows': updated_rows,
                'affected_columns': len(values)
            }
        else:
            return {
                'success': False,
                'message': f'No rows found matching conditions {conditions}'
            }
            
    except Exception as e:
        return {
            'success': False,
            'message': f'Error updating rows by conditions: {str(e)}'
        }


def delete_rows(excel_service: ExcelService, sheet_name: str = "", 
                conditions: Optional[Dict[str, Any]] = None, max_rows: Optional[int] = None, **kwargs) -> Dict[str, Any]:
    """
    Wrapper function for delete_rows operation that can be called by template registry.
    
    Args:
        excel_service: Excel service instance
        sheet_name: Target sheet name
        conditions: Conditions for which rows to delete
        max_rows: Maximum number of rows to delete (safety limit)
        **kwargs: Additional parameters
        
    Returns:
        Dict with operation result
    """
    from safety.safety_manager import SafetyManager
    
    # Create safety manager instance
    safety_manager = SafetyManager()
    
    # Create data deletion handler (we'll need to implement this)
    # For now, return a placeholder implementation
    try:
        if not excel_service.workbook:
            return {
                'success': False,
                'message': 'No Excel file loaded'
            }
        
        sheet = excel_service.get_sheet(sheet_name)
        if not sheet:
            return {
                'success': False,
                'message': f'Sheet "{sheet_name}" not found'
            }
        
        # Safety check
        if max_rows is None:
            max_rows = 10  # Default safety limit
        
        # Get sheet headers for condition matching
        headers = []
        for col in range(1, sheet.max_column + 1):
            header = sheet.cell(row=1, column=col).value
            if header:
                headers.append(str(header))
        
        deleted_count = 0
        rows_to_delete = []
        
        # Find rows to delete based on conditions
        for row_num in range(2, sheet.max_row + 1):  # Skip header row
            if deleted_count >= max_rows:
                break
            
            should_delete = False
            
            if conditions is None or not conditions:
                # No conditions specified - check if row is empty
                row_empty = True
                for col in range(1, sheet.max_column + 1):
                    if sheet.cell(row=row_num, column=col).value is not None:
                        row_empty = False
                        break
                should_delete = row_empty
            else:
                # Check if row matches all conditions
                should_delete = True
                for condition_column, condition_value in conditions.items():
                    if condition_column in headers:
                        col_index = headers.index(condition_column) + 1
                        cell_value = sheet.cell(row=row_num, column=col_index).value
                        
                        # Convert cell value to string for comparison
                        cell_str = str(cell_value) if cell_value is not None else ""
                        condition_str = str(condition_value)
                        
                        if cell_str != condition_str:
                            should_delete = False
                            break
                    else:
                        # Column not found, can't match condition
                        should_delete = False
                        break
            
            if should_delete:
                rows_to_delete.append(row_num)
                deleted_count += 1
        
        # Delete rows (in reverse order to maintain row numbers)
        for row_num in reversed(rows_to_delete):
            sheet.delete_rows(row_num)
        
        if deleted_count > 0:
            # Save workbook
            excel_service.save_workbook()
            
            condition_desc = "empty rows" if not conditions else f"rows matching {conditions}"
            return {
                'success': True,
                'message': f'Deleted {deleted_count} {condition_desc}',
                'data': {'deleted_rows': rows_to_delete, 'conditions': conditions},
                'affected_rows': deleted_count
            }
        else:
            condition_desc = "empty rows" if not conditions else f"rows matching {conditions}"
            return {
                'success': True,
                'message': f'No {condition_desc} found',
                'affected_rows': 0
            }
        
    except Exception as e:
        return {
            'success': False,
            'message': f'Error in delete_rows: {str(e)}'
        }