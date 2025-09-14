"""
Query operations for filtering, aggregating, and sorting data in Excel files.
"""

import logging
from typing import Dict, List, Optional, Any, Union
from dataclasses import dataclass

try:
    from openpyxl import Workbook
    from openpyxl.worksheet.worksheet import Worksheet
    from openpyxl.utils import get_column_letter, column_index_from_string
except ImportError:
    raise ImportError("openpyxl is required. Install with: pip install openpyxl")


@dataclass
class QueryResult:
    """Result of a query operation."""
    success: bool
    message: str
    data: Optional[List[Dict[str, Any]]] = None
    total_rows: int = 0
    filtered_rows: int = 0


class QueryOperations:
    """Handles data query operations."""
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
    
    def filter_data(
        self, 
        workbook: Workbook, 
        sheet_name: str, 
        conditions: Dict[str, Any], 
        columns: Optional[List[str]] = None
    ) -> Dict[str, Any]:
        """
        Filter data based on specified conditions.
        
        Args:
            workbook: Excel workbook
            sheet_name: Name of the sheet to filter
            conditions: Dictionary of column conditions
            columns: Specific columns to return (optional)
            
        Returns:
            Dict with filtered data
        """
        try:
            if sheet_name not in workbook.sheetnames:
                return {
                    "success": False,
                    "message": f"Sheet '{sheet_name}' not found"
                }
            
            sheet = workbook[sheet_name]
            
            # Get headers from first row
            headers = []
            for cell in sheet[1]:
                if cell.value:
                    headers.append(str(cell.value))
                else:
                    break
            
            if not headers:
                return {
                    "success": False,
                    "message": "No headers found in sheet"
                }
            
            # Extract all data
            all_data = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if any(cell is not None for cell in row):
                    row_data = {}
                    for i, value in enumerate(row):
                        if i < len(headers):
                            row_data[headers[i]] = value
                    all_data.append(row_data)
            
            # Apply filters
            filtered_data = []
            for row_data in all_data:
                matches = True
                
                for column, condition in conditions.items():
                    if column not in row_data:
                        matches = False
                        break
                    
                    value = row_data[column]
                    
                    if isinstance(condition, dict):
                        operator = condition.get("operator", "=")
                        target_value = condition.get("value")
                        
                        if not self._evaluate_condition(value, operator, target_value):
                            matches = False
                            break
                    else:
                        if value != condition:
                            matches = False
                            break
                
                if matches:
                    # Select specific columns if requested
                    if columns:
                        filtered_row = {col: row_data.get(col) for col in columns if col in row_data}
                        filtered_data.append(filtered_row)
                    else:
                        filtered_data.append(row_data)
            
            return {
                "success": True,
                "message": f"Filtered {len(filtered_data)} rows from {len(all_data)} total rows",
                "data": filtered_data,
                "total_rows": len(all_data),
                "filtered_rows": len(filtered_data)
            }
            
        except Exception as e:
            self.logger.error(f"Error filtering data: {str(e)}")
            return {
                "success": False,
                "message": f"Error filtering data: {str(e)}"
            }
    
    def aggregate_data(
        self, 
        workbook: Workbook, 
        sheet_name: str, 
        columns: List[str], 
        operation: str, 
        group_by: Optional[str] = None
    ) -> Dict[str, Any]:
        """
        Perform aggregation operations on data.
        
        Args:
            workbook: Excel workbook
            sheet_name: Name of the sheet
            columns: Columns to aggregate
            operation: Aggregation operation (sum, avg, count, max, min)
            group_by: Column to group by (optional)
            
        Returns:
            Dict with aggregated data
        """
        try:
            if sheet_name not in workbook.sheetnames:
                return {
                    "success": False,
                    "message": f"Sheet '{sheet_name}' not found"
                }
            
            sheet = workbook[sheet_name]
            
            # Get headers from first row
            headers = []
            for cell in sheet[1]:
                if cell.value:
                    headers.append(str(cell.value))
                else:
                    break
            
            # Validate columns exist
            for col in columns:
                if col not in headers:
                    return {
                        "success": False,
                        "message": f"Column '{col}' not found in sheet"
                    }
            
            if group_by and group_by not in headers:
                return {
                    "success": False,
                    "message": f"Group by column '{group_by}' not found in sheet"
                }
            
            # Extract data
            all_data = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if any(cell is not None for cell in row):
                    row_data = {}
                    for i, value in enumerate(row):
                        if i < len(headers):
                            row_data[headers[i]] = value
                    all_data.append(row_data)
            
            # Perform aggregation
            if group_by:
                # Group by specified column
                groups = {}
                for row_data in all_data:
                    group_key = row_data.get(group_by)
                    if group_key not in groups:
                        groups[group_key] = []
                    groups[group_key].append(row_data)
                
                results = {}
                for group_key, group_data in groups.items():
                    group_results = {}
                    for col in columns:
                        values = [row.get(col) for row in group_data if isinstance(row.get(col), (int, float))]
                        if values:
                            group_results[col] = self._apply_aggregation(values, operation)
                        else:
                            group_results[col] = None
                    results[group_key] = group_results
                
                return {
                    "success": True,
                    "message": f"Aggregated data by {group_by}",
                    "data": results,
                    "operation": operation,
                    "group_by": group_by
                }
            else:
                # Aggregate all data
                results = {}
                for col in columns:
                    values = [row.get(col) for row in all_data if isinstance(row.get(col), (int, float))]
                    if values:
                        results[col] = self._apply_aggregation(values, operation)
                    else:
                        results[col] = None
                
                return {
                    "success": True,
                    "message": f"Aggregated data using {operation}",
                    "data": results,
                    "operation": operation
                }
            
        except Exception as e:
            self.logger.error(f"Error aggregating data: {str(e)}")
            return {
                "success": False,
                "message": f"Error aggregating data: {str(e)}"
            }
    
    def sort_data(
        self, 
        workbook: Workbook, 
        sheet_name: str, 
        columns: List[str], 
        order: str = "asc"
    ) -> Dict[str, Any]:
        """
        Sort data by specified columns and order.
        
        Args:
            workbook: Excel workbook
            sheet_name: Name of the sheet
            columns: Columns to sort by
            order: Sort order ('asc' or 'desc')
            
        Returns:
            Dict with sorted data
        """
        try:
            if sheet_name not in workbook.sheetnames:
                return {
                    "success": False,
                    "message": f"Sheet '{sheet_name}' not found"
                }
            
            sheet = workbook[sheet_name]
            
            # Get headers from first row
            headers = []
            for cell in sheet[1]:
                if cell.value:
                    headers.append(str(cell.value))
                else:
                    break
            
            # Validate columns exist
            for col in columns:
                if col not in headers:
                    return {
                        "success": False,
                        "message": f"Column '{col}' not found in sheet"
                    }
            
            # Extract data with row numbers for sorting
            all_data = []
            for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
                if any(cell is not None for cell in row):
                    row_data = {"_row_num": row_num}
                    for i, value in enumerate(row):
                        if i < len(headers):
                            row_data[headers[i]] = value
                    all_data.append(row_data)
            
            # Sort data
            reverse_order = order.lower() == "desc"
            
            def sort_key(row):
                key_values = []
                for col in columns:
                    value = row.get(col)
                    # Handle None values and different types
                    if value is None:
                        key_values.append((0, ""))  # Sort None values first
                    elif isinstance(value, (int, float)):
                        key_values.append((1, value))
                    else:
                        key_values.append((2, str(value).lower()))
                return key_values
            
            sorted_data = sorted(all_data, key=sort_key, reverse=reverse_order)
            
            # Update the sheet with sorted data
            for i, row_data in enumerate(sorted_data, 2):
                for j, header in enumerate(headers, 1):
                    if header in row_data:
                        sheet.cell(row=i, column=j, value=row_data[header])
            
            # Remove row numbers from result data
            result_data = []
            for row_data in sorted_data:
                clean_row = {k: v for k, v in row_data.items() if k != "_row_num"}
                result_data.append(clean_row)
            
            return {
                "success": True,
                "message": f"Sorted {len(result_data)} rows by {', '.join(columns)} ({order})",
                "data": result_data,
                "sort_columns": columns,
                "sort_order": order
            }
            
        except Exception as e:
            self.logger.error(f"Error sorting data: {str(e)}")
            return {
                "success": False,
                "message": f"Error sorting data: {str(e)}"
            }
    
    def _evaluate_condition(self, value: Any, operator: str, target_value: Any) -> bool:
        """Evaluate a condition against a value."""
        try:
            if operator == "=":
                return value == target_value
            elif operator == "!=":
                return value != target_value
            elif operator == ">":
                return value > target_value
            elif operator == ">=":
                return value >= target_value
            elif operator == "<":
                return value < target_value
            elif operator == "<=":
                return value <= target_value
            elif operator == "contains":
                return str(target_value).lower() in str(value).lower()
            elif operator == "starts_with":
                return str(value).lower().startswith(str(target_value).lower())
            elif operator == "ends_with":
                return str(value).lower().endswith(str(target_value).lower())
            else:
                return False
        except Exception:
            return False
    
    def _apply_aggregation(self, values: List[Union[int, float]], operation: str) -> Optional[Union[int, float]]:
        """Apply aggregation operation to a list of values."""
        if not values:
            return None
        
        try:
            if operation.lower() == "sum":
                return sum(values)
            elif operation.lower() == "avg" or operation.lower() == "average":
                return sum(values) / len(values)
            elif operation.lower() == "count":
                return len(values)
            elif operation.lower() == "max":
                return max(values)
            elif operation.lower() == "min":
                return min(values)
            else:
                return None
        except Exception:
            return None


# Global instance for use by template system
query_operations = QueryOperations()


# Wrapper functions for template system compatibility
def filter_data(workbook, sheet_name: str, conditions: Dict[str, Any], columns: Optional[List[str]] = None) -> Dict[str, Any]:
    """Wrapper function for data filtering."""
    return query_operations.filter_data(workbook, sheet_name, conditions, columns)


def aggregate_data(workbook, sheet_name: str, columns: List[str], agg_operation: str, group_by: Optional[str] = None) -> Dict[str, Any]:
    """Wrapper function for data aggregation."""
    return query_operations.aggregate_data(workbook, sheet_name, columns, agg_operation, group_by)


def sort_data(workbook, sheet_name: str, columns: List[str], order: str = "asc") -> Dict[str, Any]:
    """Wrapper function for data sorting."""
    return query_operations.sort_data(workbook, sheet_name, columns, order)