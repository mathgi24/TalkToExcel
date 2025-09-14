"""
Excel Service for handling Excel file operations with automatic backup functionality.
"""

import os
import shutil
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Any, Tuple
from dataclasses import dataclass
import logging

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.worksheet.worksheet import Worksheet
    from openpyxl.utils import get_column_letter
except ImportError:
    raise ImportError("openpyxl is required. Install with: pip install openpyxl")


@dataclass
class ExcelStructure:
    """Data model for Excel file structure analysis."""
    sheets: List[str]
    headers: Dict[str, List[str]]
    data_types: Dict[str, Dict[str, str]]
    row_count: Dict[str, int]
    column_count: Dict[str, int]
    
    def get_sheet_info(self, sheet_name: str) -> Dict[str, Any]:
        """Get comprehensive information about a specific sheet."""
        return {
            'headers': self.headers.get(sheet_name, []),
            'data_types': self.data_types.get(sheet_name, {}),
            'row_count': self.row_count.get(sheet_name, 0),
            'column_count': self.column_count.get(sheet_name, 0)
        }


@dataclass
class BackupInfo:
    """Information about a backup file."""
    file_path: str
    timestamp: datetime
    original_file: str
    size_bytes: int


class ExcelService:
    """
    Excel service class for handling Excel file operations with automatic backup functionality.
    
    Features:
    - Automatic backup creation before every operation
    - Workbook structure analysis (headers, data types, sheet detection)
    - Backup management with timestamped files and retention policy
    - Support for .xlsx, .xls, and .csv files
    """
    
    def __init__(self, backup_dir: str = "./backups", max_backups: int = 10):
        """
        Initialize Excel service.
        
        Args:
            backup_dir: Directory to store backup files
            max_backups: Maximum number of backups to retain per file
        """
        self.backup_dir = Path(backup_dir)
        self.max_backups = max_backups
        self.workbook: Optional[Workbook] = None
        self.file_path: Optional[str] = None
        self.structure: Optional[ExcelStructure] = None
        
        # Create backup directory if it doesn't exist
        self.backup_dir.mkdir(parents=True, exist_ok=True)
        
        # Setup logging
        self.logger = logging.getLogger(__name__)
    
    def load_workbook(self, file_path: str) -> bool:
        """
        Load Excel file and analyze its structure with enhanced error handling.
        
        Args:
            file_path: Path to the Excel file
            
        Returns:
            bool: True if successful, False otherwise
        """
        try:
            # Validate file path
            if not file_path:
                raise ValueError("File path cannot be empty")
            
            file_path_obj = Path(file_path)
            
            # Check if file exists
            if not file_path_obj.exists():
                # Try to find file in common locations
                alternative_paths = [
                    Path.cwd() / file_path_obj.name,
                    Path.home() / "Documents" / file_path_obj.name,
                    Path.home() / "Desktop" / file_path_obj.name
                ]
                
                found_path = None
                for alt_path in alternative_paths:
                    if alt_path.exists():
                        found_path = alt_path
                        break
                
                if found_path:
                    raise FileNotFoundError(
                        f"File not found at '{file_path}', but found at '{found_path}'. "
                        f"Please use the correct path."
                    )
                else:
                    raise FileNotFoundError(f"File not found: {file_path}")
            
            # Check file permissions
            if not os.access(file_path, os.R_OK):
                raise PermissionError(f"No read permission for file: {file_path}")
            
            # Check file size (warn for very large files)
            file_size = file_path_obj.stat().st_size
            if file_size > 100 * 1024 * 1024:  # 100MB
                self.logger.warning(f"Large file detected ({file_size / 1024 / 1024:.1f}MB). Loading may take time.")
            
            # Check file extension
            valid_extensions = {'.xlsx', '.xls', '.xlsm'}
            if file_path_obj.suffix.lower() not in valid_extensions:
                raise ValueError(
                    f"Unsupported file format '{file_path_obj.suffix}'. "
                    f"Supported formats: {', '.join(valid_extensions)}"
                )
            
            # Close existing workbook if any
            if self.workbook:
                try:
                    self.workbook.close()
                except Exception as e:
                    self.logger.warning(f"Error closing previous workbook: {str(e)}")
            
            # Load workbook with error handling for different issues
            try:
                self.workbook = load_workbook(file_path, data_only=True)
            except PermissionError:
                raise PermissionError(
                    f"Cannot access file '{file_path}'. "
                    f"File may be open in Excel or you may lack permissions."
                )
            except Exception as e:
                error_msg = str(e).lower()
                if "corrupt" in error_msg or "invalid" in error_msg:
                    raise ValueError(
                        f"File appears to be corrupted or in an invalid format: {file_path}. "
                        f"Try opening it in Excel to repair it."
                    )
                elif "password" in error_msg or "encrypted" in error_msg:
                    raise ValueError(
                        f"File is password protected or encrypted: {file_path}. "
                        f"Please provide an unprotected version."
                    )
                else:
                    raise ValueError(f"Failed to load Excel file: {str(e)}")
            
            self.file_path = file_path
            
            # Analyze structure with error handling
            try:
                self.structure = self._analyze_structure()
            except Exception as e:
                self.logger.error(f"Failed to analyze file structure: {str(e)}")
                # Continue with basic structure
                self.structure = ExcelStructure(
                    sheets=self.workbook.sheetnames if self.workbook else [],
                    headers={},
                    data_types={},
                    row_count={},
                    column_count={}
                )
            
            self.logger.info(f"Successfully loaded workbook: {file_path}")
            return True
            
        except (FileNotFoundError, PermissionError, ValueError) as e:
            # These are expected errors that should be handled by the caller
            self.logger.error(f"Failed to load workbook {file_path}: {str(e)}")
            raise
            
        except Exception as e:
            # Unexpected errors
            self.logger.error(f"Unexpected error loading workbook {file_path}: {str(e)}")
            raise ValueError(f"Unexpected error loading Excel file: {str(e)}")
    
    def create_backup(self) -> Optional[str]:
        """
        Create a timestamped backup of the current Excel file.
        
        Returns:
            str: Path to the backup file if successful, None otherwise
        """
        if not self.file_path or not os.path.exists(self.file_path):
            self.logger.error("No file loaded or file doesn't exist")
            return None
        
        try:
            # Generate backup filename with timestamp
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            original_name = Path(self.file_path).stem
            extension = Path(self.file_path).suffix
            backup_filename = f"{original_name}_backup_{timestamp}{extension}"
            backup_path = self.backup_dir / backup_filename
            
            # Copy file to backup location
            shutil.copy2(self.file_path, backup_path)
            
            self.logger.info(f"Created backup: {backup_path}")
            
            # Clean up old backups
            self._cleanup_old_backups(original_name)
            
            return str(backup_path)
            
        except Exception as e:
            self.logger.error(f"Failed to create backup: {str(e)}")
            return None
    
    def _analyze_structure(self) -> ExcelStructure:
        """
        Analyze the structure of the loaded workbook.
        
        Returns:
            ExcelStructure: Analyzed structure information
        """
        if not self.workbook:
            raise ValueError("No workbook loaded")
        
        sheets = []
        headers = {}
        data_types = {}
        row_count = {}
        column_count = {}
        
        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            sheets.append(sheet_name)
            
            # Analyze sheet structure
            sheet_headers, sheet_data_types, rows, cols = self._analyze_sheet(sheet)
            
            headers[sheet_name] = sheet_headers
            data_types[sheet_name] = sheet_data_types
            row_count[sheet_name] = rows
            column_count[sheet_name] = cols
        
        return ExcelStructure(
            sheets=sheets,
            headers=headers,
            data_types=data_types,
            row_count=row_count,
            column_count=column_count
        )
    
    def _analyze_sheet(self, sheet: Worksheet) -> Tuple[List[str], Dict[str, str], int, int]:
        """
        Analyze a single worksheet structure.
        
        Args:
            sheet: Worksheet to analyze
            
        Returns:
            Tuple of (headers, data_types, row_count, column_count)
        """
        headers = []
        data_types = {}
        
        # Get dimensions
        max_row = sheet.max_row or 0
        max_col = sheet.max_column or 0
        
        # Handle completely empty sheets
        if max_row == 0 or max_col == 0:
            return headers, data_types, 0, 0
        
        # Handle sheets with only empty cells
        has_data = False
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                if sheet.cell(row=row, column=col).value is not None:
                    has_data = True
                    break
            if has_data:
                break
        
        if not has_data:
            return headers, data_types, 0, 0
        
        # Extract headers from first row
        for col in range(1, max_col + 1):
            cell_value = sheet.cell(row=1, column=col).value
            header = str(cell_value) if cell_value is not None else f"Column_{col}"
            headers.append(header)
        
        # Analyze data types by sampling first few data rows
        sample_rows = min(10, max_row - 1)  # Sample up to 10 rows (excluding header)
        
        for col_idx, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_idx)
            detected_type = self._detect_column_type(sheet, col_idx, 2, sample_rows + 1)
            data_types[header] = detected_type
        
        return headers, data_types, max_row, max_col
    
    def _detect_column_type(self, sheet: Worksheet, col: int, start_row: int, end_row: int) -> str:
        """
        Detect the data type of a column by sampling values.
        
        Args:
            sheet: Worksheet to analyze
            col: Column number (1-based)
            start_row: Starting row for sampling
            end_row: Ending row for sampling
            
        Returns:
            str: Detected data type ('text', 'number', 'date', 'boolean', 'mixed')
        """
        types_found = set()
        
        for row in range(start_row, min(end_row + 1, sheet.max_row + 1)):
            cell_value = sheet.cell(row=row, column=col).value
            
            if cell_value is None:
                continue
            
            if isinstance(cell_value, bool):
                types_found.add('boolean')
            elif isinstance(cell_value, (int, float)):
                types_found.add('number')
            elif isinstance(cell_value, datetime):
                types_found.add('date')
            else:
                types_found.add('text')
        
        if len(types_found) == 0:
            return 'empty'
        elif len(types_found) == 1:
            return list(types_found)[0]
        else:
            return 'mixed'
    
    def get_backup_list(self, original_filename: Optional[str] = None) -> List[BackupInfo]:
        """
        Get list of available backups.
        
        Args:
            original_filename: Filter backups for specific file (without extension)
            
        Returns:
            List[BackupInfo]: List of backup information
        """
        backups = []
        
        try:
            for backup_file in self.backup_dir.glob("*_backup_*"):
                if backup_file.is_file():
                    # Parse backup filename
                    parts = backup_file.stem.split('_backup_')
                    if len(parts) == 2:
                        orig_name = parts[0]
                        timestamp_str = parts[1]
                        
                        # Filter by original filename if specified
                        if original_filename and orig_name != original_filename:
                            continue
                        
                        try:
                            timestamp = datetime.strptime(timestamp_str, "%Y%m%d_%H%M%S")
                            size = backup_file.stat().st_size
                            
                            backup_info = BackupInfo(
                                file_path=str(backup_file),
                                timestamp=timestamp,
                                original_file=orig_name,
                                size_bytes=size
                            )
                            backups.append(backup_info)
                        except ValueError:
                            # Skip files with invalid timestamp format
                            continue
            
            # Sort by timestamp (newest first)
            backups.sort(key=lambda x: x.timestamp, reverse=True)
            
        except Exception as e:
            self.logger.error(f"Failed to get backup list: {str(e)}")
        
        return backups
    
    def restore_from_backup(self, backup_path: str) -> bool:
        """
        Restore Excel file from a backup.
        
        Args:
            backup_path: Path to the backup file
            
        Returns:
            bool: True if successful, False otherwise
        """
        try:
            if not os.path.exists(backup_path):
                raise FileNotFoundError(f"Backup file not found: {backup_path}")
            
            if not self.file_path:
                raise ValueError("No original file path set")
            
            # Copy backup to original location
            shutil.copy2(backup_path, self.file_path)
            
            # Reload the workbook
            self.load_workbook(self.file_path)
            
            self.logger.info(f"Successfully restored from backup: {backup_path}")
            return True
            
        except Exception as e:
            self.logger.error(f"Failed to restore from backup {backup_path}: {str(e)}")
            return False
    
    def _cleanup_old_backups(self, original_filename: str) -> None:
        """
        Clean up old backup files, keeping only the most recent ones.
        
        Args:
            original_filename: Original filename (without extension) to clean up
        """
        try:
            backups = self.get_backup_list(original_filename)
            
            if len(backups) > self.max_backups:
                # Remove oldest backups
                backups_to_remove = backups[self.max_backups:]
                
                for backup in backups_to_remove:
                    try:
                        os.remove(backup.file_path)
                        self.logger.info(f"Removed old backup: {backup.file_path}")
                    except Exception as e:
                        self.logger.warning(f"Failed to remove backup {backup.file_path}: {str(e)}")
        
        except Exception as e:
            self.logger.error(f"Failed to cleanup old backups: {str(e)}")
    
    def save_workbook(self, create_backup: bool = True) -> bool:
        """
        Save the current workbook to file with enhanced error handling.
        
        Args:
            create_backup: Whether to create a backup before saving
            
        Returns:
            bool: True if successful, False otherwise
        """
        if not self.workbook or not self.file_path:
            self.logger.error("No workbook loaded or file path not set")
            return False
        
        backup_path = None
        
        try:
            # Create backup before saving if requested
            if create_backup:
                backup_path = self.create_backup()
                if not backup_path:
                    self.logger.warning("Failed to create backup, but continuing with save")
            
            # Check write permissions before attempting save
            file_dir = os.path.dirname(self.file_path) or '.'
            if not os.access(file_dir, os.W_OK):
                raise PermissionError(f"No write permission for directory: {file_dir}")
            
            # Check if file is locked (Windows specific check)
            if os.path.exists(self.file_path):
                try:
                    # Try to open file in write mode to check if it's locked
                    with open(self.file_path, 'r+b'):
                        pass
                except (PermissionError, OSError) as e:
                    if "being used by another process" in str(e) or "Permission denied" in str(e):
                        raise PermissionError(
                            f"File is currently open in Excel or another application: {self.file_path}. "
                            f"Please close the file and try again."
                        )
                    else:
                        raise
            
            # Save workbook with error handling
            try:
                self.workbook.save(self.file_path)
            except PermissionError as e:
                raise PermissionError(
                    f"Cannot save file - it may be open in Excel or you may lack permissions: {self.file_path}"
                )
            except OSError as e:
                if "No space left" in str(e):
                    raise OSError(f"Insufficient disk space to save file: {self.file_path}")
                else:
                    raise OSError(f"System error while saving file: {str(e)}")
            
            self.logger.info(f"Successfully saved workbook: {self.file_path}")
            return True
            
        except (PermissionError, OSError) as e:
            self.logger.error(f"Failed to save workbook: {str(e)}")
            
            # If we have a backup and save failed, offer to restore
            if backup_path and os.path.exists(backup_path):
                self.logger.info(f"Save failed, backup available at: {backup_path}")
            
            raise
            
        except Exception as e:
            self.logger.error(f"Unexpected error saving workbook: {str(e)}")
            
            # If we have a backup and save failed, offer to restore
            if backup_path and os.path.exists(backup_path):
                self.logger.info(f"Save failed, backup available at: {backup_path}")
            
            raise ValueError(f"Unexpected error saving Excel file: {str(e)}")
    
    def get_structure(self) -> Optional[ExcelStructure]:
        """
        Get the analyzed structure of the current workbook.
        
        Returns:
            ExcelStructure: Structure information or None if no workbook loaded
        """
        return self.structure
    
    def get_sheet_names(self) -> List[str]:
        """
        Get list of sheet names in the workbook.
        
        Returns:
            List[str]: Sheet names
        """
        if not self.workbook:
            return []
        return self.workbook.sheetnames
    
    def get_sheet(self, sheet_name: str) -> Optional[Worksheet]:
        """
        Get a specific worksheet by name.
        
        Args:
            sheet_name: Name of the sheet
            
        Returns:
            Worksheet: The worksheet or None if not found
        """
        if not self.workbook:
            return None
        
        try:
            return self.workbook[sheet_name]
        except KeyError:
            return None
    
    def close(self) -> None:
        """Close the workbook and clean up resources."""
        if self.workbook:
            try:
                self.workbook.close()
                self.logger.info("Workbook closed")
            except Exception as e:
                self.logger.warning(f"Error closing workbook: {e}")
            finally:
                self.workbook = None
                self.file_path = None
                self.structure = None
    
    def cleanup(self):
        """Clean up resources used by the service."""
        self.close()