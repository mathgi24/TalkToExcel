"""
Main command processor that orchestrates LLM and operations.

This module provides the central command processing pipeline that:
1. Parses natural language commands using LLM
2. Validates and routes operations through safety checks
3. Executes operations using the template registry
4. Provides user feedback and error handling
"""

import logging
from typing import Dict, Any, Optional, List
from dataclasses import dataclass
from enum import Enum

from llm.ollama_service import OllamaService, LLMResponse, OllamaConnectionError
from templates.template_registry import TemplateRegistry
from safety.safety_manager import SafetyManager, SafetyResult
from excel.excel_service import ExcelService
from operations.crud_handlers import DataInsertionHandler, DataQueryHandler
from operations.visualization_operations import VisualizationOperations

# Add src directory to Python path for imports
import sys
from pathlib import Path
current_dir = Path(__file__).parent
src_dir = current_dir.parent if current_dir.name != 'src' else current_dir
if str(src_dir) not in sys.path:
    sys.path.insert(0, str(src_dir))



class ProcessingStatus(Enum):
    """Status of command processing."""
    SUCCESS = "success"
    FAILED = "failed"
    BLOCKED = "blocked"
    CONFIRMATION_REQUIRED = "confirmation_required"
    CLARIFICATION_NEEDED = "clarification_needed"


@dataclass
class ProcessingResult:
    """Result of command processing."""
    status: ProcessingStatus
    message: str
    data: Optional[Any] = None
    confirmation_prompt: Optional[str] = None
    clarification_questions: Optional[List[str]] = None
    operation_details: Optional[Dict[str, Any]] = None
    safety_report: Optional[str] = None
    warnings: Optional[List[str]] = None
    
    def __post_init__(self):
        if self.warnings is None:
            self.warnings = []


class CommandProcessor:
    """
    Main command processor that orchestrates LLM and operations.
    
    This class serves as the central hub for processing natural language commands
    by coordinating between the LLM service, safety systems, and operation handlers.
    """
    
    def __init__(self, 
                 llm_service: OllamaService,
                 template_registry: TemplateRegistry,
                 safety_manager: SafetyManager,
                 excel_service: ExcelService):
        """
        Initialize command processor with required services.
        
        Args:
            llm_service: LLM service for natural language processing
            template_registry: Registry for operation templates
            safety_manager: Safety manager for validation and risk assessment
            excel_service: Excel service for file operations
        """
        self.llm_service = llm_service
        self.template_registry = template_registry
        self.safety_manager = safety_manager
        self.excel_service = excel_service
        
        # Initialize operation handlers
        self.data_insertion_handler = DataInsertionHandler(excel_service, safety_manager)
        self.data_query_handler = DataQueryHandler(excel_service, safety_manager)
        self.visualization_operations = VisualizationOperations()
        
        self.logger = logging.getLogger(__name__)
        
        # Track pending confirmations
        self._pending_confirmations: Dict[str, Dict[str, Any]] = {}
    
    def process_command_with_fields(self, original_command: str, sheet_name: str, 
                                   category_field: str, value_field: str) -> ProcessingResult:
        """Process a chart command with specific field selections."""
        try:
            # Create a mock LLM response with the specific field parameters
            from llm.ollama_service import LLMResponse
            
            # Find the column indices for the specified fields
            sheet = self.excel_service.get_sheet(sheet_name)
            if not sheet:
                return ProcessingResult(
                    status=ProcessingStatus.FAILED,
                    message=f"Sheet '{sheet_name}' not found"
                )
            
            # Get headers and find column positions
            headers = {}
            for col in range(1, sheet.max_column + 1):
                header = sheet.cell(row=1, column=col).value
                if header:
                    headers[str(header).lower()] = col
            
            category_col = headers.get(category_field.lower())
            value_col = headers.get(value_field.lower())
            
            if not category_col or not value_col:
                available_fields = list(headers.keys())
                return ProcessingResult(
                    status=ProcessingStatus.FAILED,
                    message=f"Could not find fields '{category_field}' or '{value_field}'. Available fields: {', '.join(available_fields)}"
                )
            
            # Create data range using the specific columns
            from openpyxl.utils import get_column_letter
            category_col_letter = get_column_letter(category_col)
            value_col_letter = get_column_letter(value_col)
            
            # Create a range that includes both columns
            start_col = min(category_col, value_col)
            end_col = max(category_col, value_col)
            data_range = f"{get_column_letter(start_col)}1:{get_column_letter(end_col)}{sheet.max_row}"
            
            # Create LLM response with specific parameters
            llm_response = LLMResponse(
                intent="visualization_operations",
                operation="create_chart",
                parameters={
                    "sheet_name": sheet_name,
                    "data_range": data_range,
                    "chart_type": "pie",
                    "title": f"{sheet_name} - {category_field} vs {value_field}",
                    "category_field": category_field,
                    "value_field": value_field
                },
                confidence=0.9,
                reasoning=f"User specified {category_field} for categories and {value_field} for values"
            )
            
            # Process normally with the specific parameters
            safety_result = self._evaluate_safety(llm_response, original_command)
            
            if safety_result.blocked:
                return ProcessingResult(
                    status=ProcessingStatus.BLOCKED,
                    message=f"Operation blocked for safety: {'; '.join(safety_result.error_messages)}"
                )
            
            return self._execute_operation(llm_response, safety_result)
            
        except Exception as e:
            self.logger.error(f"Error processing command with fields: {str(e)}")
            return ProcessingResult(
                status=ProcessingStatus.FAILED,
                message=f"Error processing chart with specified fields: {str(e)}"
            )

    def process_command(self, user_command: str, 
                       confirmation_id: Optional[str] = None,
                       user_confirmed: Optional[bool] = None) -> ProcessingResult:
        """
        Process a natural language command through the complete pipeline.
        
        Args:
            user_command: Natural language command from user
            confirmation_id: ID of pending confirmation (if responding to confirmation)
            user_confirmed: User's confirmation response (True/False)
            
        Returns:
            ProcessingResult: Complete processing result with status and data
        """
        try:
            # Store the original user command for use in operations
            self._current_user_command = user_command
            
            # Handle confirmation responses
            if confirmation_id and confirmation_id in self._pending_confirmations:
                return self._handle_confirmation_response(confirmation_id, user_confirmed)
            
            # Step 1: Parse command using LLM
            self.logger.info(f"Processing command: {user_command}")
            llm_response = self._parse_command(user_command)
            
            if not llm_response:
                return ProcessingResult(
                    status=ProcessingStatus.FAILED,
                    message="Failed to parse command. Please try rephrasing your request."
                )
            
            # Step 2: Handle clarification requests
            if llm_response.intent == "clarification_needed":
                return self._handle_clarification_needed(llm_response, user_command)
            
            # Step 3: Validate command structure
            validation_result = self._validate_command_structure(llm_response)
            if not validation_result.status == ProcessingStatus.SUCCESS:
                return validation_result
            
            # Step 4: Safety evaluation
            safety_result = self._evaluate_safety(llm_response, user_command)
            
            # Step 5: Handle blocked operations
            if safety_result.blocked:
                return ProcessingResult(
                    status=ProcessingStatus.BLOCKED,
                    message=f"Operation blocked for safety: {'; '.join(safety_result.error_messages)}",
                    safety_report=self.safety_manager.get_detailed_report(safety_result),
                    warnings=safety_result.suggestions
                )
            
            # Step 6: Handle confirmation requirements
            if safety_result.confirmation_required:
                return self._handle_confirmation_required(llm_response, safety_result)
            
            # Step 7: Route and execute operation
            return self._execute_operation(llm_response, safety_result)
            
        except Exception as e:
            self.logger.error(f"Error processing command: {str(e)}")
            return ProcessingResult(
                status=ProcessingStatus.FAILED,
                message=f"An error occurred while processing your command: {str(e)}"
            )
    
    def _parse_command(self, user_command: str) -> Optional[LLMResponse]:
        """Parse natural language command using LLM service."""
        try:
            return self.llm_service.parse_to_structured_command(user_command)
        except OllamaConnectionError as e:
            self.logger.error(f"LLM connection error: {str(e)}")
            return None
        except Exception as e:
            self.logger.error(f"Error parsing command: {str(e)}")
            return None
    
    def _validate_command_structure(self, llm_response: LLMResponse) -> ProcessingResult:
        """Validate the structure of the parsed command."""
        # Check if LLM response is valid
        if not self.llm_service.validate_response(llm_response):
            return ProcessingResult(
                status=ProcessingStatus.FAILED,
                message="Invalid command structure. Please try rephrasing your request."
            )
        
        # Check if operation exists in template registry
        if not self._is_operation_available(llm_response.intent, llm_response.operation):
            # Get available operations for the specified intent
            available_ops = self.template_registry.get_operations_by_category(llm_response.intent)
            
            # Also check if the operation exists in other categories
            all_operations = self.template_registry.get_all_operations()
            matching_ops = [op for op in all_operations if op.endswith(f".{llm_response.operation}")]
            
            if matching_ops:
                return ProcessingResult(
                    status=ProcessingStatus.FAILED,
                    message=f"Operation '{llm_response.operation}' not available in '{llm_response.intent}'. Found in: {', '.join(matching_ops)}. Available in '{llm_response.intent}': {', '.join(available_ops)}"
                )
            else:
                return ProcessingResult(
                    status=ProcessingStatus.FAILED,
                    message=f"Operation '{llm_response.operation}' not available. Available operations in '{llm_response.intent}': {', '.join(available_ops)}"
                )
        
        # Special validation for chart operations - check if we need field clarification
        if (llm_response.intent == "visualization_operations" and 
            llm_response.operation == "create_chart" and 
            llm_response.parameters.get("chart_type") == "pie"):
            
            clarification_result = self._check_chart_field_clarification(llm_response)
            if clarification_result:
                return clarification_result
        
        return ProcessingResult(status=ProcessingStatus.SUCCESS, message="Command structure valid")
    
    def _check_chart_field_clarification(self, llm_response: LLMResponse) -> Optional[ProcessingResult]:
        """Check if chart creation needs field clarification."""
        sheet_name = llm_response.parameters.get('sheet_name')
        data_range = llm_response.parameters.get('data_range')
        
        # If data_range is just a sheet name (auto-detect mode), we should ask for specific fields
        if (sheet_name and data_range and 
            ':' not in data_range and 
            data_range in self.excel_service.get_sheet_names()):
            
            # Get available columns from the sheet
            sheet = self.excel_service.get_sheet(sheet_name)
            if sheet and sheet.max_row > 1:
                # Get column headers
                headers = []
                for col in range(1, sheet.max_column + 1):
                    header = sheet.cell(row=1, column=col).value
                    if header:
                        headers.append(str(header))
                
                if len(headers) > 2:  # If there are multiple columns, ask for clarification
                    # Analyze column types to suggest categories and values
                    text_columns = []
                    numeric_columns = []
                    
                    for col_idx, header in enumerate(headers, 1):
                        # Check sample data to determine column type
                        sample_value = sheet.cell(row=2, column=col_idx).value
                        if sample_value is not None:
                            if isinstance(sample_value, str):
                                text_columns.append(header)
                            elif isinstance(sample_value, (int, float)):
                                numeric_columns.append(header)
                    
                    # Generate clarification questions
                    clarification_questions = [
                        f"I found these columns in the '{sheet_name}' sheet: {', '.join(headers)}",
                        f"For a pie chart, I need:",
                        f"• A category column (text): {', '.join(text_columns) if text_columns else 'No text columns found'}",
                        f"• A value column (numbers): {', '.join(numeric_columns) if numeric_columns else 'No numeric columns found'}",
                        f"Please specify which columns to use. For example: 'Use {text_columns[0] if text_columns else headers[0]} for categories and {numeric_columns[0] if numeric_columns else headers[1]} for values'"
                    ]
                    
                    return ProcessingResult(
                        status=ProcessingStatus.CLARIFICATION_NEEDED,
                        message="I need to know which specific columns to use for the pie chart.",
                        clarification_questions=clarification_questions
                    )
        
        return None
    
    def _evaluate_safety(self, llm_response: LLMResponse, user_command: str) -> SafetyResult:
        """Evaluate safety of the operation."""
        # Get sheet information if available
        sheet_info = None
        if self.excel_service.get_structure():
            target_sheet = llm_response.parameters.get('sheet_name') or llm_response.parameters.get('target_sheet')
            if target_sheet:
                sheet_info = self.excel_service.get_structure().get_sheet_info(target_sheet)
        
        return self.safety_manager.evaluate_operation(
            operation=llm_response.operation,  # Use just the operation name, not intent.operation
            parameters=llm_response.parameters,
            command_text=user_command,
            sheet_info=sheet_info
        )
    
    def _handle_clarification_needed(self, llm_response: LLMResponse, user_command: str) -> ProcessingResult:
        """Handle cases where LLM needs clarification."""
        clarification_questions = []
        
        # Generate clarification questions based on the error
        error_info = llm_response.parameters.get('error', '')
        
        if 'ambiguous' in error_info.lower():
            clarification_questions.append("Could you be more specific about which data you want to work with?")
            clarification_questions.append("Which sheet should I use for this operation?")
        elif 'unclear' in error_info.lower():
            clarification_questions.append("Could you rephrase your request with more details?")
            clarification_questions.append("What specific action would you like me to perform?")
        else:
            clarification_questions.append("I didn't understand your request. Could you provide more details?")
            clarification_questions.append("What would you like me to do with your Excel file?")
        
        return ProcessingResult(
            status=ProcessingStatus.CLARIFICATION_NEEDED,
            message="I need more information to process your request.",
            clarification_questions=clarification_questions
        )
    
    def _handle_confirmation_required(self, llm_response: LLMResponse, safety_result: SafetyResult) -> ProcessingResult:
        """Handle operations that require user confirmation."""
        # Generate unique confirmation ID
        import uuid
        confirmation_id = str(uuid.uuid4())
        
        # Store pending confirmation
        self._pending_confirmations[confirmation_id] = {
            'llm_response': llm_response,
            'safety_result': safety_result,
            'timestamp': self._get_current_timestamp()
        }
        
        # Generate confirmation prompt
        confirmation_prompt = self.safety_manager.create_confirmation_prompt(safety_result)
        
        return ProcessingResult(
            status=ProcessingStatus.CONFIRMATION_REQUIRED,
            message="This operation requires confirmation due to safety considerations.",
            confirmation_prompt=confirmation_prompt,
            operation_details={
                'confirmation_id': confirmation_id,
                'operation': f"{llm_response.intent}.{llm_response.operation}",
                'parameters': llm_response.parameters,
                'risk_level': safety_result.risk_assessment.level.value
            },
            safety_report=self.safety_manager.get_detailed_report(safety_result)
        )
    
    def _handle_confirmation_response(self, confirmation_id: str, user_confirmed: Optional[bool]) -> ProcessingResult:
        """Handle user's response to confirmation prompt."""
        if confirmation_id not in self._pending_confirmations:
            return ProcessingResult(
                status=ProcessingStatus.FAILED,
                message="Confirmation request not found or expired."
            )
        
        pending = self._pending_confirmations[confirmation_id]
        llm_response = pending['llm_response']
        safety_result = pending['safety_result']
        
        # Remove from pending confirmations
        del self._pending_confirmations[confirmation_id]
        
        if user_confirmed is None or not user_confirmed:
            return ProcessingResult(
                status=ProcessingStatus.FAILED,
                message="Operation cancelled by user."
            )
        
        # User confirmed - proceed with execution
        return self._execute_operation(llm_response, safety_result)
    
    def _execute_operation(self, llm_response: LLMResponse, safety_result: SafetyResult) -> ProcessingResult:
        """Execute the validated and approved operation."""
        try:
            # Route to appropriate operation handler
            result = self._route_operation(llm_response)
            
            if result:
                return ProcessingResult(
                    status=ProcessingStatus.SUCCESS,
                    message=result.get('message', 'Operation completed successfully'),
                    data=result.get('data'),
                    operation_details={
                        'operation': f"{llm_response.intent}.{llm_response.operation}",
                        'parameters': llm_response.parameters,
                        'execution_details': result
                    },
                    warnings=safety_result.warnings if safety_result.warnings else None
                )
            else:
                return ProcessingResult(
                    status=ProcessingStatus.FAILED,
                    message="Operation execution failed"
                )
                
        except Exception as e:
            self.logger.error(f"Error executing operation: {str(e)}")
            return ProcessingResult(
                status=ProcessingStatus.FAILED,
                message=f"Operation execution failed: {str(e)}"
            )
    
    def _route_operation(self, llm_response: LLMResponse) -> Optional[Dict[str, Any]]:
        """Route operation to appropriate handler based on intent."""
        intent = llm_response.intent
        operation = llm_response.operation
        parameters = llm_response.parameters
        
        try:
            # Check if operation is available in the specified intent
            actual_intent = intent
            if not self.template_registry.is_operation_available(intent, operation):
                # Look for the operation in other categories
                all_operations = self.template_registry.get_all_operations()
                for op_key in all_operations:
                    if op_key.endswith(f".{operation}"):
                        actual_intent = op_key.split('.')[0]
                        self.logger.info(f"Redirecting '{intent}.{operation}' to '{actual_intent}.{operation}'")
                        break
            
            # Route to template registry for execution
            if self.template_registry.is_operation_available(actual_intent, operation):
                # Add excel_service and original command to parameters
                enhanced_parameters = parameters.copy()
                enhanced_parameters['excel_service'] = self.excel_service
                enhanced_parameters['original_command'] = getattr(self, '_current_user_command', '')
                
                # Auto-fill and validate sheet_name
                available_sheets = self.excel_service.get_sheet_names()
                if 'sheet_name' not in enhanced_parameters or not enhanced_parameters['sheet_name']:
                    # Use first available sheet if none specified
                    if available_sheets:
                        enhanced_parameters['sheet_name'] = available_sheets[0]
                else:
                    # Try to match partial sheet names to full names
                    requested_sheet = enhanced_parameters['sheet_name']
                    if requested_sheet not in available_sheets:
                        # Try case-insensitive partial matching
                        requested_lower = requested_sheet.lower()
                        matched_sheet = None
                        
                        # First try exact case-insensitive match
                        for sheet in available_sheets:
                            if sheet.lower() == requested_lower:
                                matched_sheet = sheet
                                break
                        
                        # Then try partial matching
                        if not matched_sheet:
                            for sheet in available_sheets:
                                if requested_lower in sheet.lower() or sheet.lower().startswith(requested_lower):
                                    matched_sheet = sheet
                                    break
                        
                        if matched_sheet:
                            enhanced_parameters['sheet_name'] = matched_sheet
                            self.logger.info(f"Matched '{requested_sheet}' to '{matched_sheet}'")
                        else:
                            self.logger.warning(f"Sheet '{requested_sheet}' not found in {available_sheets}")
                
                # Normalize conditions parameter
                if 'conditions' in enhanced_parameters:
                    conditions = enhanced_parameters['conditions']
                    if isinstance(conditions, str):
                        # Try to interpret string conditions as sheet name hints
                        if conditions.lower() in ['employee', 'employees']:
                            # This might be a hint about which sheet to use
                            for sheet in available_sheets:
                                if 'employee' in sheet.lower():
                                    enhanced_parameters['sheet_name'] = sheet
                                    enhanced_parameters['conditions'] = None
                                    break
                        else:
                            # For other string conditions, set to None for now
                            enhanced_parameters['conditions'] = None
                
                # Add default limit for query operations if missing
                if intent == "data_operations" and operation == "query_data":
                    if 'limit' not in enhanced_parameters or enhanced_parameters['limit'] is None:
                        enhanced_parameters['limit'] = 100
                        self.logger.info("Added default limit of 100 for query operation")
                
                # Filter parameters for visualization operations
                if intent == "visualization_operations":
                    # Remove excel_service and original_command parameters for visualization operations
                    filtered_parameters = {k: v for k, v in enhanced_parameters.items() 
                                         if k not in ['excel_service', 'original_command']}
                    # Add workbook parameter instead
                    filtered_parameters['workbook'] = self.excel_service.workbook
                    enhanced_parameters = filtered_parameters
                
                # Execute through template registry
                result = self.template_registry.execute_operation(actual_intent, operation, **enhanced_parameters)
                
                # Format result for consistent return structure
                if hasattr(result, 'success'):
                    # Handle OperationResult objects
                    return {
                        'success': result.success,
                        'message': result.message,
                        'data': getattr(result, 'data', None),
                        'affected_rows': getattr(result, 'affected_rows', 0),
                        'affected_columns': getattr(result, 'affected_columns', 0)
                    }
                elif isinstance(result, dict):
                    # Handle dictionary results
                    return result
                else:
                    # Handle other result types
                    return {
                        'success': True,
                        'message': 'Operation completed',
                        'data': result
                    }
            else:
                # Fallback to direct operation routing
                return self._direct_operation_routing(intent, operation, parameters)
                
        except NotImplementedError as e:
            self.logger.warning(f"Operation not implemented: {str(e)}")
            return {
                'success': False,
                'message': f"Operation '{intent}.{operation}' is not yet implemented"
            }
        except Exception as e:
            self.logger.error(f"Error routing operation: {str(e)}")
            return {
                'success': False,
                'message': f"Error executing operation: {str(e)}"
            }
    
    def _direct_operation_routing(self, intent: str, operation: str, parameters: Dict[str, Any]) -> Optional[Dict[str, Any]]:
        """Direct routing for operations not handled by template registry."""
        # Handle CRUD operations
        if intent == "data_operations":
            if operation == "create_data":
                # Route to data insertion handler
                from operations.crud_handlers import InsertionData
                
                # Handle both dictionary and list data formats
                data_param = parameters.get('data', [])
                target_sheet = parameters.get('sheet_name', '')
                
                # Convert dictionary to list in correct column order
                if isinstance(data_param, dict):
                    # Get sheet headers to determine column order
                    if target_sheet and self.excel_service.workbook:
                        try:
                            sheet = self.excel_service.get_sheet(target_sheet)
                            if sheet and sheet.max_row > 0:
                                # Get headers from first row
                                headers = []
                                for col in range(1, sheet.max_column + 1):
                                    header = sheet.cell(row=1, column=col).value
                                    if header:
                                        headers.append(str(header))
                                
                                # Convert dict to list using header order
                                values_list = []
                                for header in headers:
                                    if header in data_param:
                                        values_list.append(data_param[header])
                                    else:
                                        values_list.append(None)  # Empty cell for missing data
                                
                                data_values = values_list
                            else:
                                # No headers found, use dict values in order
                                data_values = list(data_param.values())
                        except Exception as e:
                            self.logger.warning(f"Error getting sheet headers: {e}")
                            # Fallback to dict values
                            data_values = list(data_param.values())
                    else:
                        # No sheet specified or workbook not loaded, use dict values
                        data_values = list(data_param.values())
                elif isinstance(data_param, list):
                    data_values = data_param
                else:
                    data_values = [data_param] if data_param else []
                
                insertion_data = InsertionData(
                    values=data_values,
                    target_sheet=target_sheet,
                    target_row=parameters.get('target_row'),
                    target_column=parameters.get('target_column'),
                    column_names=parameters.get('column_names')
                )
                result = self.data_insertion_handler.insert_row(insertion_data)
                return {
                    'success': result.success,
                    'message': result.message,
                    'data': result.data,
                    'affected_rows': result.affected_rows
                }
            
            elif operation == "query_data":
                # Route to data query handler
                from operations.crud_handlers import QueryData
                query_data = QueryData(
                    target_sheet=parameters.get('sheet_name', ''),
                    columns=parameters.get('columns'),
                    conditions=parameters.get('conditions'),
                    sort_by=parameters.get('sort_by'),
                    limit=parameters.get('limit', 100)
                )
                result = self.data_query_handler.query_data(query_data)
                return {
                    'success': result.success,
                    'message': result.message,
                    'data': result.data,
                    'row_count': result.row_count
                }
        
        # Handle visualization operations
        elif intent == "visualization_operations":
            if operation == "create_chart":
                if not self.excel_service.workbook:
                    return {
                        'success': False,
                        'message': 'No Excel file loaded'
                    }
                
                result = self.visualization_operations.create_chart(
                    workbook=self.excel_service.workbook,
                    sheet_name=parameters.get('sheet_name', ''),
                    data_range=parameters.get('data_range', ''),
                    chart_type=parameters.get('chart_type'),
                    title=parameters.get('title'),
                    excel_service=self.excel_service,
                    category_field=parameters.get('category_field'),
                    value_field=parameters.get('value_field')
                )
                return result
        
        # Handle data analysis operations
        elif intent == "data_analysis_operations":
            if not self.excel_service.workbook:
                return {
                    'success': False,
                    'message': 'No Excel file loaded'
                }
            
            if operation == "calculate_statistics":
                from operations.data_analysis_operations import calculate_statistics
                result = calculate_statistics(
                    workbook=self.excel_service.workbook,
                    sheet_name=parameters.get('sheet_name', ''),
                    column=parameters.get('column', ''),
                    stat_type=parameters.get('stat_type', 'all')
                )
                return result
            
            elif operation == "find_outliers":
                from operations.data_analysis_operations import find_outliers
                result = find_outliers(
                    workbook=self.excel_service.workbook,
                    sheet_name=parameters.get('sheet_name', ''),
                    column=parameters.get('column', ''),
                    method=parameters.get('method', 'iqr')
                )
                return result
            
            elif operation == "data_summary":
                from operations.data_analysis_operations import data_summary
                result = data_summary(
                    workbook=self.excel_service.workbook,
                    sheet_name=parameters.get('sheet_name', '')
                )
                return result
        
        return None
    
    def _is_operation_available(self, intent: str, operation: str) -> bool:
        """Check if an operation is available in the template registry."""
        # First check the specified intent
        if self.template_registry.is_operation_available(intent, operation):
            return True
        
        # If not found in specified intent, check all categories
        # This handles cases where LLM picks wrong intent but correct operation
        all_operations = self.template_registry.get_all_operations()
        for op_key in all_operations:
            if op_key.endswith(f".{operation}"):
                self.logger.info(f"Operation '{operation}' found in '{op_key}' instead of '{intent}.{operation}'")
                return True
        
        return False
    
    def _get_current_timestamp(self) -> float:
        """Get current timestamp for tracking purposes."""
        import time
        return time.time()
    
    def get_available_operations(self) -> Dict[str, List[str]]:
        """Get list of available operations by category."""
        return {
            category: self.template_registry.get_operations_by_category(category)
            for category in self.template_registry.get_registry_stats()['category_names']
        }
    
    def get_operation_help(self, intent: str, operation: str) -> Optional[Dict[str, Any]]:
        """Get help information for a specific operation."""
        metadata = self.template_registry.get_operation_metadata(intent, operation)
        if metadata:
            return {
                'operation': f"{intent}.{operation}",
                'config': metadata.get('config', {}),
                'parameters': metadata.get('config', {}).get('parameters', []),
                'examples': metadata.get('config', {}).get('examples', []),
                'safety_level': metadata.get('config', {}).get('safety_level', 'unknown')
            }
        return None
    
    def cleanup_expired_confirmations(self, max_age_seconds: int = 300) -> None:
        """Clean up expired confirmation requests."""
        current_time = self._get_current_timestamp()
        expired_ids = []
        
        for conf_id, conf_data in self._pending_confirmations.items():
            if current_time - conf_data['timestamp'] > max_age_seconds:
                expired_ids.append(conf_id)
        
        for conf_id in expired_ids:
            del self._pending_confirmations[conf_id]
        
        if expired_ids:
            self.logger.info(f"Cleaned up {len(expired_ids)} expired confirmations")